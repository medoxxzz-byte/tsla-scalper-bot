"""
Smart Trading Alert Bot - V8.0 (Mosquito ثاقب + Options Scalper)
Webhook Server for TSLA Mosquito Swamp Pine Script

NEW IN V8.0:
  - محرك تداول أوبشن تلقائي بالكامل (Options Scalper)
  - طبقتين: ATM Scalp (سكالبينج سريع) + ITM Pullback (عقد عميق)
  - 0DTE فقط — الاثنين للخميس
  - نافذة التداول: 10:10 AM - 12:40 PM ET
  - TP/SL تلقائي + تعزيز ذكي
  - حد خسارة محفظة: $7,000
  - اكتشاف ترند تلقائي (EMA + VWAP + ADX + Momentum)
  - خريطة انعكاسات للحماية من المناطق الخطيرة
  - اسم البوت: ثاقب V8 (Options Scalper)

KEPT FROM V7.0:
  - خريطة الانعكاسات المدمجة مع رسائل التلقرام (Reversal Map Alignment)
  - فلتر Trend Alignment (يمنع صفقات ضد الترند)
  - تقليل Reversal Cooldown من 30 إلى 15 دقيقة
  - فلتر ADX إجباري (لا تداول في سوق جانبي ADX < 20)
  - Stop Loss تلقائي -20% (محاكاة على السهم)
  - FlashAlpha GEX كفلتر إضافي
  - نافذة تداول: 9:30 AM - 3:30 PM ET
  - Alpaca Paper Trading Auto-Executor
  - Position tracker with real-time P&L

الهدف: تجربة 3-6 أشهر لإثبات الاستراتيجية قبل التداول الحقيقي.

V8 Architecture: options_scalper.py runs as autonomous background thread.
"""

import os
import json
import time
import logging
import threading
from datetime import datetime, timezone, timedelta

try:
    from flask import Flask, request, jsonify, render_template
except ImportError:
    os.system("pip install flask")
    from flask import Flask, request, jsonify, render_template

try:
    import requests as http_requests
except ImportError:
    os.system("pip install requests")
    import requests as http_requests

try:
    import yfinance as yf
except ImportError:
    os.system("pip install yfinance")
    import yfinance as yf

try:
    import pandas as pd
except ImportError:
    os.system("pip install pandas")
    import pandas as pd

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

## ── FlashAlpha GEX Integration (V7.1 — Direct API) ─────────────────────────
FLASHALPHA_API_KEY = os.environ.get("FLASHALPHA_API_KEY", "srd7RXM1awDGPt6XkSuPpxHVnJD2XQsUu8UeYUZJ")
FLASHALPHA_BASE_URL = "https://lab.flashalpha.com/v1"
_FA_AVAILABLE = True  # Always available — direct API

# GEX cache — يتحدث مرتين باليوم (9:25 AM + 10:00 AM ET)
_gex_cache = {
    "data": None,
    "last_fetch": None,
    "fetch_count_today": 0,
    "fetch_date": ""
}
_GEX_MAX_DAILY_FETCHES = 3  # حد الطلبات اليومية (من 5 المجانية)


def fetch_flashalpha_gex():
    """سحب بيانات GEX من FlashAlpha API (Free tier: 5 requests/day)."""
    global _gex_cache
    try:
        today = get_today()
    except:
        today = (datetime.now(timezone.utc) - timedelta(hours=4)).strftime("%Y-%m-%d")
    
    # Reset counter for new day
    if _gex_cache["fetch_date"] != today:
        _gex_cache["fetch_count_today"] = 0
        _gex_cache["fetch_date"] = today
    
    # Check daily limit
    if _gex_cache["fetch_count_today"] >= _GEX_MAX_DAILY_FETCHES:
        logging.info(f"[GEX] Daily fetch limit reached ({_GEX_MAX_DAILY_FETCHES}). Using cache.")
        return _gex_cache["data"]
    
    try:
        headers = {"X-Api-Key": FLASHALPHA_API_KEY}
        r = http_requests.get(
            f"{FLASHALPHA_BASE_URL}/exposure/levels/TSLA",
            headers=headers,
            timeout=15
        )
        if r.status_code == 200:
            data = r.json()
            levels = data.get("levels", {})
            gex_data = {
                "gamma_flip": levels.get("gamma_flip"),
                "call_wall": levels.get("call_wall"),
                "put_wall": levels.get("put_wall"),
                "max_positive_gamma": levels.get("max_positive_gamma"),
                "max_negative_gamma": levels.get("max_negative_gamma"),
                "highest_oi_strike": levels.get("highest_oi_strike"),
                "zero_dte_magnet": levels.get("zero_dte_magnet"),
                "underlying_price": data.get("underlying_price"),
                "as_of": data.get("as_of")
            }
            _gex_cache["data"] = gex_data
            _gex_cache["last_fetch"] = datetime.now(timezone.utc).isoformat()
            _gex_cache["fetch_count_today"] += 1
            logging.info(f"[GEX] FlashAlpha data fetched OK — Gamma Flip: ${gex_data['gamma_flip']:.2f}, "
                         f"Call Wall: ${gex_data['call_wall']}, Put Wall: ${gex_data['put_wall']} "
                         f"(fetch {_gex_cache['fetch_count_today']}/{_GEX_MAX_DAILY_FETCHES})")
            return gex_data
        elif r.status_code == 429:
            logging.warning(f"[GEX] FlashAlpha quota exceeded — using cache")
            return _gex_cache["data"]
        else:
            logging.error(f"[GEX] FlashAlpha error: {r.status_code} {r.text[:200]}")
            return _gex_cache["data"]
    except Exception as e:
        logging.error(f"[GEX] FlashAlpha fetch exception: {e}")
        return _gex_cache["data"]


def get_gex_levels():
    """إرجاع بيانات GEX المخزنة (أو سحبها إذا ما موجودة)."""
    if _gex_cache["data"]:
        return _gex_cache["data"]
    return fetch_flashalpha_gex()


def check_gex_alignment(signal, price):
    """فحص توافق الإشارة مع GEX."""
    gex = get_gex_levels()
    if not gex or not gex.get("gamma_flip"):
        return True, "GEX غير متاح"
    
    gamma_flip = float(gex["gamma_flip"])
    call_wall = float(gex["call_wall"]) if gex.get("call_wall") else None
    put_wall = float(gex["put_wall"]) if gex.get("put_wall") else None
    price = float(price)
    
    if signal == "CALL":
        # CALL قرب Call Wall = خطر
        if call_wall and abs(price - call_wall) / call_wall < 0.005:
            return False, f"السعر عند Call Wall ${call_wall} — مقاومة قوية"
        # CALL تحت Gamma Flip = ضد التيار
        if price < gamma_flip * 0.99:
            return True, f"⚠️ تحت Gamma Flip ${gamma_flip:.0f} — حذر"
        return True, f"✅ فوق Gamma Flip ${gamma_flip:.0f}"
    
    elif signal == "PUT":
        # PUT قرب Put Wall = خطر
        if put_wall and abs(price - put_wall) / put_wall < 0.005:
            return False, f"السعر عند Put Wall ${put_wall} — دعم قوي"
        # PUT فوق Gamma Flip = ضد التيار
        if price > gamma_flip * 1.01:
            return True, f"⚠️ فوق Gamma Flip ${gamma_flip:.0f} — حذر"
        return True, f"✅ تحت Gamma Flip ${gamma_flip:.0f}"
    
    return True, ""


def format_gex_telegram():
    """تنسيق رسالة GEX للتلقرام."""
    gex = get_gex_levels()
    if not gex:
        return None
    
    price = gex.get("underlying_price", 0)
    gamma_flip = gex.get("gamma_flip", 0)
    call_wall = gex.get("call_wall", 0)
    put_wall = gex.get("put_wall", 0)
    max_pos = gex.get("max_positive_gamma", 0)
    
    # تحديد النظام
    if price and gamma_flip:
        if float(price) > float(gamma_flip):
            regime = "إيجابي ✅ (فوق Gamma Flip)"
            regime_icon = "🟢"
        else:
            regime = "سلبي ⚠️ (تحت Gamma Flip)"
            regime_icon = "🔴"
    else:
        regime = "غير محدد"
        regime_icon = "⚪"
    
    now_et_str = (datetime.now(timezone.utc) - timedelta(hours=4)).strftime("%I:%M %p")
    
    msg = (
        f"{regime_icon} <b>خريطة GEX — TSLA</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"💰 السعر: <code>${float(price):.2f}</code>\n"
        f"📊 النظام: {regime}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🟡 Gamma Flip: <code>${float(gamma_flip):.0f}</code> ← المحور\n"
        f"🔴 Call Wall: <code>${float(call_wall):.0f}</code> ← مقاومة\n"
        f"🟢 Put Wall: <code>${float(put_wall):.0f}</code> ← دعم\n"
        f"⭐ Max +Gamma: <code>${float(max_pos):.0f}</code> ← جاذب\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🕐 {now_et_str} ET"
    )
    return msg

# ── Reversal Detector Integration ─────────────────────────────────────────
try:
    import reversal_detector as rd
    _RD_AVAILABLE = True
except ImportError:
    _RD_AVAILABLE = False
    logging.warning("reversal_detector.py not found — reversal detection disabled")

# ── Reversal Tracker Integration ─────────────────────────────────────────
try:
    import reversal_tracker as rt
    _RT_AVAILABLE = True
except ImportError:
    _RT_AVAILABLE = False
    logging.warning("reversal_tracker.py not found — reversal tracking disabled")

# ── Alpaca Options Feed Integration ─────────────────────────────────────────
try:
    import alpaca_options as ao
    _AO_AVAILABLE = True
except ImportError:
    _AO_AVAILABLE = False
    logging.warning("alpaca_options.py not found — options feed disabled")

# ── حالة الصفقة الحالية لتحديد نوع الأوبشن ────────────────────────────────────
_current_signal_type = None  # "CALL" أو "PUT"

def _get_signal_type():
    return _current_signal_type

def _get_tsla_price_for_options():
    """جلب سعر TSLA من Alpaca (بديل yfinance الموثوق)."""
    return get_tsla_price_alpaca_snapshot()

def _on_option_data_received(opt: dict):
    """استقبال بيانات الأوبشن الحية وتمريرها لـ reversal_detector."""
    if _RD_AVAILABLE:
        rd.update_option_data(
            premium      = opt["premium"],
            delta        = opt["delta"],
            option_symbol= opt["symbol"],
            strike_price = opt["strike"],
            expiration   = opt["expiration"],
            option_type  = opt["option_type"]
        )

# ──────────────────────────────────────────────────────────────────────────────
# Configuration
# ──────────────────────────────────────────────────────────────────────────────

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "8708530077:AAF16LsdHUNTW5G25UypCm8NiFTmCIranP8")
TELEGRAM_CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID",   "975644160")

SERVER_HOST    = os.environ.get("SERVER_HOST", "0.0.0.0")
SERVER_PORT    = int(os.environ.get("PORT", os.environ.get("SERVER_PORT", "8080")))
WEBHOOK_SECRET = os.environ.get("WEBHOOK_SECRET", "")

# ── Alpaca Configuration ──────────────────────────────────────────────────────
ALPACA_API_KEY    = os.environ.get("ALPACA_API_KEY",    "PKW3OHVLGGWGYCFMTCKDB435WA")
ALPACA_SECRET_KEY = os.environ.get("ALPACA_SECRET_KEY", "BeNQ9BiZ8t5wxDwb6Dmvd62W3i57wKj8SmdSTxjAQYYH")
ALPACA_BASE_URL   = os.environ.get("ALPACA_BASE_URL",   "https://paper-api.alpaca.markets")

# ── Trading Window (ET) ───────────────────────────────────────────────────────
TRADING_START_HOUR   = 9
TRADING_START_MINUTE = 30
TRADING_END_HOUR     = 15
TRADING_END_MINUTE   = 30

# ── Position Sizing ───────────────────────────────────────────────────────────
MAX_CONTRACTS_PER_TRADE = 10

# ── Strategy Flags ────────────────────────────────────────────────────────────
# V10 FIX: إيقاف استراتيجية شراء TSLA السهم (تسبب خسائر ضخمة)
STOCK_STRATEGY_ENABLED = False
# V10 FIX: إيقاف Pyramid Auto (WR=33%, P&L=-$831)
PYRAMID_AUTO_ENABLED = False
MAX_OPTION_PRICE        = 5.00
MIN_OPTION_PRICE        = 0.10
MIN_STARS_TO_EXECUTE    = 3      # V7.1: أقل عدد نجوم لتنفيذ الصفقة تلقائياً

# ── Cooldowns ─────────────────────────────────────────────────────────────────
COOLDOWN_SECONDS_SIMILAR = 1500   # 25 min between same-direction trade signals
COOLDOWN_MIN_GAP         = 30     # minimum 30s between any two alerts
REVERSAL_COOLDOWN_SECS   = 900    # 15 min between reversal alerts

# V7.1: حدود يومية محسّنة
MAX_DAILY_TRADE_ALERTS    = 6    # حد تنبيهات التداول (CALL/PUT)
MAX_DAILY_REVERSAL_ALERTS = 3    # حد تنبيهات الانعكاس
MAX_DAILY_ALERTS          = int(os.environ.get("MAX_DAILY_TRADES", "6"))

KEEP_ALIVE_INTERVAL = 600

# V6.0: فلتر ADX إجباري (لا تداول في سوق جانبي)
MIN_ADX_TO_TRADE = 20

# V6.0: Stop Loss تلقائي -20% (محاكاة على السهم)
AUTO_STOP_LOSS_PCT = 0.20

LOSS_COUNTER_MAX      = 3
LOSS_COOLDOWN_SECONDS = 1800

# V7.1: حد الخسارة اليومي
DAILY_LOSS_LIMIT = -150.0  # -$150

# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("alert_bot_v71.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Flask App & State
# ──────────────────────────────────────────────────────────────────────────────

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

def _create_retry_session():
    session = http_requests.Session()
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "POST", "DELETE"],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

_retry_session = _create_retry_session()

app = Flask(__name__)

alert_history = []
MAX_HISTORY   = 100

last_alert_time   = 0
last_alert_price  = ""
last_alert_signal = ""
last_call_time    = 0
last_put_time     = 0
last_reversal_time = 0

consecutive_no_move = 0
loss_cooldown_until = 0

daily_alerts  = []
daily_reversal_count = 0  # V7.1: عداد الانعكاسات اليومي
daily_date    = ""
blocked_today = []

market_state = {
    "last_price": 0.0,
    "last_updated": None,
    "day_high": 0.0,
    "day_low": 0.0
}

# Alpaca position tracker
active_positions = {}

# ──────────────────────────────────────────────────────────────────────────────
# V7.0: Reversal Map — خريطة الانعكاسات المدمجة مع الرسائل
# ──────────────────────────────────────────────────────────────────────────────

reversal_map = {
    "levels": [],
    "date": "",
    "built": False
}

MAP_PROXIMITY_PCT = 0.008
MAP_NEAR_PCT      = 0.015


def build_reversal_map():
    global reversal_map
    today = get_today()
    
    if reversal_map["date"] == today and reversal_map["built"] and len(reversal_map["levels"]) > 0:
        return reversal_map
    
    levels = []
    
    # ── المصدر 1: GEX Levels (FlashAlpha Direct API) ─────────────────
    if _FA_AVAILABLE:
        gex = get_gex_levels()
        if gex:
            if gex.get("call_wall"):
                levels.append({
                    "name": "Call Wall",
                    "price": float(gex["call_wall"]),
                    "type": "resistance",
                    "source": "GEX",
                    "strength": "قوية"
                })
            if gex.get("put_wall"):
                levels.append({
                    "name": "Put Wall",
                    "price": float(gex["put_wall"]),
                    "type": "support",
                    "source": "GEX",
                    "strength": "قوية"
                })
            if gex.get("gamma_flip"):
                levels.append({
                    "name": "Gamma Flip",
                    "price": float(gex["gamma_flip"]),
                    "type": "pivot",
                    "source": "GEX",
                    "strength": "قوية جداً"
                })
    
    # ── المصدر 2: Fibonacci Levels ───────────────────────────────────────────
    d_high = market_state.get("day_high", 0)
    d_low  = market_state.get("day_low", 0)
    if d_high > 0 and d_low > 0 and d_high > d_low:
        fib = calc_fibonacci(d_high, d_low)
        levels.append({"name": "Fib 38.2%", "price": fib["fib_382"], "type": "support", "source": "Fibonacci", "strength": "متوسطة"})
        levels.append({"name": "Fib 50%",   "price": fib["fib_500"], "type": "pivot",   "source": "Fibonacci", "strength": "قوية"})
        levels.append({"name": "Fib 61.8%", "price": fib["fib_618"], "type": "support", "source": "Fibonacci", "strength": "قوية"})
    
    # ── المصدر 3: S/R from 5-min candles ─────────────────────────────────────
    sr_r, sr_s = get_tsla_sr_levels()
    if sr_r and sr_r > 0:
        levels.append({"name": "5m Resistance", "price": float(sr_r), "type": "resistance", "source": "S/R", "strength": "متوسطة"})
    if sr_s and sr_s > 0:
        levels.append({"name": "5m Support", "price": float(sr_s), "type": "support", "source": "S/R", "strength": "متوسطة"})
    
    # ── المصدر 4: Round Numbers (أرقام نفسية) ────────────────────────────────
    if d_high > 0:
        mid_price = (d_high + d_low) / 2 if d_low > 0 else d_high
        base = int(mid_price / 5) * 5
        for rn in [base - 10, base - 5, base, base + 5, base + 10, base + 15]:
            if rn > 0:
                levels.append({
                    "name": f"${rn} Round",
                    "price": float(rn),
                    "type": "pivot",
                    "source": "Psychology",
                    "strength": "متوسطة" if rn % 10 != 0 else "قوية"
                })
    
    # ── إزالة المكررات وترتيب ────────────────────────────────────────────────
    unique_levels = []
    levels_sorted = sorted(levels, key=lambda x: x["price"], reverse=True)
    for lvl in levels_sorted:
        is_duplicate = False
        for existing in unique_levels:
            if abs(lvl["price"] - existing["price"]) / existing["price"] < 0.003:
                if _strength_rank(lvl["strength"]) > _strength_rank(existing["strength"]):
                    unique_levels.remove(existing)
                    unique_levels.append(lvl)
                is_duplicate = True
                break
        if not is_duplicate:
            unique_levels.append(lvl)
    
    reversal_map["levels"] = sorted(unique_levels, key=lambda x: x["price"], reverse=True)
    reversal_map["date"] = today
    reversal_map["built"] = True
    
    logger.info(f"[ReversalMap] Built with {len(reversal_map['levels'])} levels for {today}")
    return reversal_map


def _strength_rank(strength_str):
    ranks = {"ضعيفة": 1, "متوسطة": 2, "قوية": 3, "قوية جداً": 4}
    return ranks.get(strength_str, 1)


def check_map_alignment(price, signal_direction=None):
    if not reversal_map["built"] or not reversal_map["levels"]:
        return False, "", None
    
    try:
        price = float(price)
    except (ValueError, TypeError):
        return False, "", None
    
    nearest = None
    min_dist_pct = float('inf')
    
    for lvl in reversal_map["levels"]:
        dist_pct = abs(price - lvl["price"]) / lvl["price"]
        if dist_pct < min_dist_pct:
            min_dist_pct = dist_pct
            nearest = lvl
    
    if not nearest:
        return False, "", None
    
    dist_dollars = abs(price - nearest["price"])
    dist_pct_display = min_dist_pct * 100
    
    if min_dist_pct <= MAP_PROXIMITY_PCT:
        position = "عند"
        is_aligned = True
        
        if signal_direction == "BULL" and nearest["type"] in ("support", "pivot"):
            alignment_text = f"✅ متوافق — عند {nearest['name']} (${nearest['price']:.2f})"
        elif signal_direction == "BEAR" and nearest["type"] in ("resistance", "pivot"):
            alignment_text = f"✅ متوافق — عند {nearest['name']} (${nearest['price']:.2f})"
        elif signal_direction:
            alignment_text = f"⚠️ تعارض — عند {nearest['name']} (${nearest['price']:.2f})"
            is_aligned = False
        else:
            alignment_text = f"📍 عند {nearest['name']} (${nearest['price']:.2f})"
        
    elif min_dist_pct <= MAP_NEAR_PCT:
        is_aligned = True
        above_below = "فوق" if price > nearest["price"] else "تحت"
        alignment_text = f"📍 قريب من {nearest['name']} (${nearest['price']:.2f}) | {above_below}"
    else:
        is_aligned = False
        above_below = "فوق" if price > nearest["price"] else "تحت"
        alignment_text = f"📊 أقرب مستوى: {nearest['name']} (${nearest['price']:.2f}) | {above_below}"
    
    return is_aligned, alignment_text, nearest

# ──────────────────────────────────────────────────────────────────────────────
# Helper Functions
# ──────────────────────────────────────────────────────────────────────────────

def get_et_now():
    return datetime.now(timezone.utc) - timedelta(hours=4)

def get_today():
    return get_et_now().strftime("%Y-%m-%d")

def is_friday():
    """V7.1: تحقق إذا اليوم جمعة (0DTE = كازينو)."""
    return get_et_now().weekday() == 4

def is_trading_window():
    now = get_et_now()
    current_minutes = now.hour * 60 + now.minute
    start_minutes   = TRADING_START_HOUR * 60 + TRADING_START_MINUTE
    end_minutes     = TRADING_END_HOUR   * 60 + TRADING_END_MINUTE

    if now.weekday() >= 5:
        return False, "عطلة نهاية الأسبوع"

    if current_minutes < start_minutes:
        remaining = start_minutes - current_minutes
        return False, f"قبل نافذة التداول — يبدأ 9:30 AM ET (بعد {remaining} دقيقة)"

    if current_minutes >= end_minutes:
        return False, "انتهت نافذة التداول — 3:30 PM ET"

    return True, "نافذة التداول مفتوحة (9:30-3:30) ✅"

def reset_daily_if_needed():
    global daily_date, daily_alerts, blocked_today, daily_reversal_count
    global consecutive_no_move, loss_cooldown_until, last_reversal_time
    today = get_today()
    if daily_date != today:
        daily_date = today
        daily_alerts = []
        blocked_today = []
        daily_reversal_count = 0
        consecutive_no_move = 0
        loss_cooldown_until = 0
        last_reversal_time = 0
        logger.info(f"--- V7.1 Reset daily limits for {today} ---")

def safe_get(data, key, default=""):
    val = data.get(key, default)
    return str(val).strip() if val is not None else default

# ──────────────────────────────────────────────────────────────────────────────
# Alpaca API Functions
# ──────────────────────────────────────────────────────────────────────────────

def alpaca_headers():
    return {
        "APCA-API-KEY-ID":     ALPACA_API_KEY,
        "APCA-API-SECRET-KEY": ALPACA_SECRET_KEY,
        "Content-Type":        "application/json"
    }

def get_alpaca_account():
    try:
        r = _retry_session.get(
            f"{ALPACA_BASE_URL}/v2/account",
            headers=alpaca_headers(),
            timeout=10
        )
        if r.status_code == 200:
            return r.json()
        logger.error(f"Alpaca account error: {r.status_code} {r.text}")
        return None
    except Exception as e:
        logger.error(f"Alpaca account exception: {e}")
        return None

def get_alpaca_positions():
    try:
        r = _retry_session.get(
            f"{ALPACA_BASE_URL}/v2/positions",
            headers=alpaca_headers(),
            timeout=10
        )
        if r.status_code == 200:
            return r.json()
        return []
    except Exception as e:
        logger.error(f"Alpaca positions error: {e}")
        return []

def place_alpaca_stock_order(symbol, qty, side, order_type="market",
                              limit_price=None, stop_price=None,
                              time_in_force="day"):
    payload = {
        "symbol":        symbol,
        "qty":           str(qty),
        "side":          side,
        "type":          order_type,
        "time_in_force": time_in_force,
    }
    if order_type == "limit" and limit_price:
        payload["limit_price"] = str(round(limit_price, 2))
    if order_type == "stop" and stop_price:
        payload["stop_price"] = str(round(stop_price, 2))

    try:
        r = _retry_session.post(
            f"{ALPACA_BASE_URL}/v2/orders",
            headers=alpaca_headers(),
            json=payload,
            timeout=15
        )
        if r.status_code in (200, 201):
            order = r.json()
            logger.info(f"Alpaca order placed: {order.get('id')} | {side} {qty} {symbol}")
            return order
        else:
            logger.error(f"Alpaca order error: {r.status_code} {r.text}")
            return None
    except Exception as e:
        logger.error(f"Alpaca order exception: {e}")
        return None

def close_alpaca_position(symbol):
    try:
        r = _retry_session.delete(
            f"{ALPACA_BASE_URL}/v2/positions/{symbol}",
            headers=alpaca_headers(),
            timeout=10
        )
        if r.status_code in (200, 204):
            logger.info(f"Alpaca position closed: {symbol}")
            return True
        logger.error(f"Close position error: {r.status_code} {r.text}")
        return False
    except Exception as e:
        logger.error(f"Close position exception: {e}")
        return False

def execute_trade(signal, price, stars, opt_data):
    """
    Execute a trade on Alpaca Paper Trading.
    CALL → BUY TSLA shares | PUT → SELL SHORT TSLA shares
    """
    # V10 FIX: استراتيجية السهم موقفة بسبب خسائر ضخمة (-$9,500)
    if not STOCK_STRATEGY_ENABLED:
        logger.info(f"[execute_trade] DISABLED — STOCK_STRATEGY_ENABLED=False ({signal})")
        return False, None, "⛔ استراتيجية السهم موقفة"
    try:
        account = get_alpaca_account()
        if not account:
            return False, None, "فشل الاتصال بـ Alpaca"

        buying_power = float(account.get("buying_power", 0))
        tsla_price = float(price)
        qty = MAX_CONTRACTS_PER_TRADE

        if signal == "CALL":
            if buying_power < tsla_price * qty:
                return False, None, f"رصيد غير كافٍ (${buying_power:.0f})"

            order = place_alpaca_stock_order("TSLA", qty, "buy")
            if not order:
                return False, None, "فشل تنفيذ أمر الشراء"

            order_id = order.get("id", "")
            tp_stock = tsla_price * 1.025
            sl_stock = tsla_price * (1 - AUTO_STOP_LOSS_PCT * 0.075)

            active_positions[order_id] = {
                "signal":     signal,
                "tsla_price": tsla_price,
                "qty":        qty,
                "tp":         round(tp_stock, 2),
                "sl":         round(sl_stock, 2),
                "entry_time": get_et_now().strftime("%I:%M %p"),
                "stars":      stars,
                "opt_data":   opt_data
            }

            msg = (
                f"🤖 <b>ثاقب نفّذ — CALL</b>\n"
                f"━━━━━━━━━━━━━━━\n"
                f"🟢 شراء TSLA × {qty} سهم @ ${tsla_price:.2f}\n"
                f"🎯 هدف: ${tp_stock:.2f}\n"
                f"🛑 وقف: ${sl_stock:.2f}\n"
                f"⭐ {stars}/5\n"
                f"🕐 {get_et_now().strftime('%I:%M %p')} ET"
            )
            return True, order_id, msg

        elif signal == "PUT":
            if account.get("shorting_enabled", False):
                order = place_alpaca_stock_order("TSLA", qty, "sell")
                if not order:
                    return False, None, "فشل تنفيذ أمر البيع"

                order_id = order.get("id", "")
                tp_stock = tsla_price * 0.985
                sl_stock = tsla_price * 1.010

                active_positions[order_id] = {
                    "signal":     signal,
                    "tsla_price": tsla_price,
                    "qty":        qty,
                    "tp":         round(tp_stock, 2),
                    "sl":         round(sl_stock, 2),
                    "entry_time": get_et_now().strftime("%I:%M %p"),
                    "stars":      stars,
                    "opt_data":   opt_data
                }

                msg = (
                    f"🤖 <b>ثاقب نفّذ — PUT</b>\n"
                    f"━━━━━━━━━━━━━━━\n"
                    f"🔴 بيع (Short) TSLA × {qty} سهم @ ${tsla_price:.2f}\n"
                    f"🎯 هدف: ${tp_stock:.2f}\n"
                    f"🛑 وقف: ${sl_stock:.2f}\n"
                    f"⭐ {stars}/5\n"
                    f"🕐 {get_et_now().strftime('%I:%M %p')} ET"
                )
                return True, order_id, msg
            else:
                fake_id = f"PAPER_{int(time.time())}"
                tp_stock = tsla_price * 0.985
                sl_stock = tsla_price * 1.010

                active_positions[fake_id] = {
                    "signal":     signal,
                    "tsla_price": tsla_price,
                    "qty":        qty,
                    "tp":         round(tp_stock, 2),
                    "sl":         round(sl_stock, 2),
                    "entry_time": get_et_now().strftime("%I:%M %p"),
                    "stars":      stars,
                    "opt_data":   opt_data,
                    "paper_only": True
                }

                msg = (
                    f"🤖 <b>ثاقب سجّل — PUT (محاكاة)</b>\n"
                    f"━━━━━━━━━━━━━━━\n"
                    f"🔴 PUT TSLA × {qty} @ ${tsla_price:.2f}\n"
                    f"🎯 هدف: ${tp_stock:.2f}\n"
                    f"🛑 وقف: ${sl_stock:.2f}\n"
                    f"⭐ {stars}/5\n"
                    f"ℹ️ البيع القصير غير مفعل — تسجيل فقط\n"
                    f"🕐 {get_et_now().strftime('%I:%M %p')} ET"
                )
                return True, fake_id, msg

    except Exception as e:
        logger.error(f"execute_trade error: {e}")
        return False, None, f"خطأ في التنفيذ: {e}"

# ──────────────────────────────────────────────────────────────────────────────
# Position Monitor (background thread)
# ──────────────────────────────────────────────────────────────────────────────

def monitor_positions():
    while True:
        time.sleep(60)
        if not active_positions:
            continue

        try:
            current_price = get_tsla_price_alpaca_snapshot()
            if current_price <= 0:
                continue

            now = get_et_now()
            positions_to_close = []

            for order_id, pos in list(active_positions.items()):
                signal     = pos["signal"]
                entry      = pos["tsla_price"]
                tp         = pos["tp"]
                sl         = pos["sl"]
                entry_time = pos["entry_time"]
                stars      = pos["stars"]
                paper_only = pos.get("paper_only", False)

                hit_tp = False
                hit_sl = False
                close_reason = ""
                pnl = 0.0

                if signal == "CALL":
                    pnl = (current_price - entry) * pos["qty"]
                    if current_price >= tp:
                        hit_tp = True
                        close_reason = "🎯 وصل الهدf"
                    elif current_price <= sl:
                        hit_sl = True
                        close_reason = "🛑 وقف الخسارة"
                elif signal == "PUT":
                    pnl = (entry - current_price) * pos["qty"]
                    if current_price <= tp:
                        hit_tp = True
                        close_reason = "🎯 وصل الهدf"
                    elif current_price >= sl:
                        hit_sl = True
                        close_reason = "🛑 وقف الخسارة"

                force_close = (now.hour == 15 and now.minute >= 45)

                if hit_tp or hit_sl or force_close:
                    if force_close:
                        close_reason = "⏰ إغلاق إجباري (قبل إغلاق السوق)"

                    if not paper_only:
                        close_alpaca_position("TSLA")

                    pnl_icon = "💚" if pnl >= 0 else "🔴"
                    msg = (
                        f"{pnl_icon} <b>ثاقب أغلق الصفقة</b>\n"
                        f"━━━━━━━━━━━━━━━\n"
                        f"{'🟢 CALL' if signal == 'CALL' else '🔴 PUT'} TSLA\n"
                        f"📥 دخول: ${entry:.2f} ({entry_time})\n"
                        f"📤 خروج: ${current_price:.2f}\n"
                        f"{close_reason}\n"
                        f"{'💰' if pnl >= 0 else '💸'} <b>P&L: {'+' if pnl >= 0 else ''}{pnl:.2f}$</b>\n"
                        f"🕐 {now.strftime('%I:%M %p')} ET"
                    )
                    send_telegram(msg)
                    positions_to_close.append(order_id)
                    logger.info(f"Position closed: {order_id} | P&L: {pnl:.2f} | Reason: {close_reason}")

            for oid in positions_to_close:
                active_positions.pop(oid, None)

        except Exception as e:
            logger.error(f"monitor_positions error: {e}")

# ──────────────────────────────────────────────────────────────────────────────
# Fibonacci Calculator
# ──────────────────────────────────────────────────────────────────────────────

def calc_fibonacci(high, low):
    diff = high - low
    return {
        "high":    round(high, 2),
        "low":     round(low, 2),
        "fib_236": round(high - diff * 0.236, 2),
        "fib_382": round(high - diff * 0.382, 2),
        "fib_500": round(high - diff * 0.500, 2),
        "fib_618": round(high - diff * 0.618, 2),
        "fib_786": round(high - diff * 0.786, 2),
    }

def get_tsla_price_alpaca():
    try:
        r = _retry_session.get(
            f"{ALPACA_BASE_URL}/v2/stocks/TSLA/quotes/latest",
            headers=alpaca_headers(),
            timeout=8
        )
        if r.status_code == 200:
            data = r.json()
            price = float(data.get("quote", {}).get("ap", 0) or data.get("quote", {}).get("bp", 0))
            if price > 0:
                return price
    except Exception as e:
        logger.error(f"Alpaca price error: {e}")
    return 0.0

def get_tsla_price_alpaca_snapshot():
    try:
        r = _retry_session.get(
            "https://data.alpaca.markets/v2/stocks/TSLA/snapshot",
            headers=alpaca_headers(),
            timeout=8
        )
        if r.status_code == 200:
            snap = r.json()
            price = float(snap.get("latestTrade", {}).get("p", 0))
            if price > 0:
                return price
            price = float(snap.get("latestQuote", {}).get("ap", 0))
            if price > 0:
                return price
    except Exception as e:
        logger.error(f"Alpaca snapshot error: {e}")
    return get_tsla_price_alpaca()

def get_tsla_day_data_alpaca():
    try:
        r = _retry_session.get(
            "https://data.alpaca.markets/v2/stocks/TSLA/snapshot",
            headers=alpaca_headers(),
            timeout=8
        )
        if r.status_code == 200:
            snap = r.json()
            daily = snap.get("dailyBar", {})
            latest = snap.get("latestTrade", {})
            last_price = float(latest.get("p", 0))
            day_high   = float(daily.get("h", 0))
            day_low    = float(daily.get("l", 0))
            if last_price > 0:
                return last_price, day_high, day_low
    except Exception as e:
        logger.error(f"Alpaca day data error: {e}")
    return None, None, None

def get_tsla_5min_bars_alpaca():
    try:
        from datetime import datetime, timedelta
        end = datetime.utcnow()
        start = end - timedelta(hours=8)
        params = {
            "timeframe": "5Min",
            "start": start.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "end": end.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "limit": 10,
            "feed": "iex"
        }
        r = http_requests.get(
            "https://data.alpaca.markets/v2/stocks/TSLA/bars",
            headers=alpaca_headers(),
            params=params,
            timeout=10
        )
        if r.status_code == 200:
            bars = r.json().get("bars", [])
            if len(bars) >= 4:
                return bars
    except Exception as e:
        logger.error(f"Alpaca 5min bars error: {e}")
    return []

def get_tsla_day_data():
    price, high, low = get_tsla_day_data_alpaca()
    if price and price > 0:
        return price, high or 0.0, low or 0.0
    try:
        tkr = yf.Ticker("TSLA")
        info = tkr.fast_info
        day_high   = float(info.day_high)   if info.day_high   else 0.0
        day_low    = float(info.day_low)    if info.day_low    else 0.0
        last_price = float(info.last_price) if info.last_price else 0.0
        if last_price > 0:
            return last_price, day_high, day_low
    except Exception as e:
        logger.error(f"yfinance day data error: {e}")
    return None, None, None

# ──────────────────────────────────────────────────────────────────────────────
# Support & Resistance — 5-Minute Candles
# ──────────────────────────────────────────────────────────────────────────────

_sr_cache = {
    "resistance":    None,
    "support":       None,
    "last_candle_ts": None,
}

def get_tsla_sr_levels():
    global _sr_cache
    try:
        bars = get_tsla_5min_bars_alpaca()
        if bars and len(bars) >= 4:
            closed = bars[-4:-1]
            latest_ts = closed[-1].get("t", "")
            if _sr_cache["last_candle_ts"] is not None and latest_ts == _sr_cache["last_candle_ts"]:
                return _sr_cache["resistance"], _sr_cache["support"]
            resistance = round(max(b["h"] for b in closed), 2)
            support    = round(min(b["l"] for b in closed), 2)
            _sr_cache["resistance"]     = resistance
            _sr_cache["support"]        = support
            _sr_cache["last_candle_ts"] = latest_ts
            return resistance, support
    except Exception as e:
        logger.error(f"[S&R] Alpaca bars error: {e}")
    try:
        tkr = yf.Ticker("TSLA")
        df = tkr.history(period="1d", interval="5m")
        if df is not None and len(df) >= 4:
            closed_candles = df.iloc[-4:-1]
            latest_candle_ts = closed_candles.index[-1]
            if _sr_cache["last_candle_ts"] is not None and latest_candle_ts == _sr_cache["last_candle_ts"]:
                return _sr_cache["resistance"], _sr_cache["support"]
            resistance = round(float(closed_candles["High"].max()), 2)
            support    = round(float(closed_candles["Low"].min()),  2)
            _sr_cache["resistance"]     = resistance
            _sr_cache["support"]        = support
            _sr_cache["last_candle_ts"] = latest_candle_ts
            return resistance, support
    except Exception as e:
        logger.error(f"[S&R] yfinance error: {e}")
    return _sr_cache["resistance"], _sr_cache["support"]

# ──────────────────────────────────────────────────────────────────────────────
# Options Data (Yahoo Finance)
# ──────────────────────────────────────────────────────────────────────────────

def get_best_option(ticker_symbol, signal_type, current_price):
    try:
        current_price = float(current_price)
        try:
            tkr = yf.Ticker(ticker_symbol)
            exp_dates = tkr.options
        except Exception as e:
            logger.error(f"yfinance options error: {e}")
            return None, None
        if not exp_dates:
            return None, None

        today_str = get_today()
        valid_dates = [d for d in exp_dates if d >= today_str]
        if not valid_dates:
            return None, None

        primary_expiry = valid_dates[0]
        alt_expiry     = valid_dates[1] if len(valid_dates) > 1 else None

        primary_opt = _find_contract(tkr, primary_expiry, signal_type, current_price)
        alt_opt     = _find_contract(tkr, alt_expiry, signal_type, current_price) if alt_expiry else None

        return primary_opt, alt_opt

    except Exception as e:
        logger.error(f"Error fetching option: {e}")
        return None, None

def _find_contract(tkr, expiry, signal_type, current_price):
    try:
        opt_chain = tkr.option_chain(expiry)
        df = opt_chain.calls if signal_type == "CALL" else opt_chain.puts
        if df.empty:
            return None

        df['strike_diff'] = abs(df['strike'] - current_price)

        if signal_type == "CALL":
            valid_strikes = df[df['strike'] <= current_price * 1.005]
        else:
            valid_strikes = df[df['strike'] >= current_price * 0.995]

        if valid_strikes.empty:
            valid_strikes = df

        best_row   = valid_strikes.loc[valid_strikes['strike_diff'].idxmin()]
        last_price = float(best_row.get('lastPrice', 0))
        ask        = float(best_row.get('ask', 0))
        bid        = float(best_row.get('bid', 0))

        if last_price <= 0 and ask > 0:
            last_price = ask

        tp = last_price * 1.40
        sl = last_price * 0.50

        is_0dte = (expiry == get_today())

        return {
            "strike":             float(best_row['strike']),
            "expiry":             expiry,
            "symbol":             best_row['contractSymbol'],
            "last_price":         last_price,
            "ask":                ask,
            "bid":                bid,
            "volume":             int(best_row.get('volume', 0)),
            "open_interest":      int(best_row.get('openInterest', 0)),
            "implied_volatility": float(best_row.get('impliedVolatility', 0)),
            "tp":                 tp,
            "sl":                 sl,
            "is_0dte":            is_0dte
        }
    except Exception as e:
        logger.error(f"Error finding contract for {expiry}: {e}")
        return None

# ──────────────────────────────────────────────────────────────────────────────
# V7.1: Stars Rating — FIXED
# ──────────────────────────────────────────────────────────────────────────────

def calculate_stars_v71(data):
    """
    V7.1: إعادة حساب النجوم في الباك إند.
    
    Pine Script يرسل النجوم الأصلية (1-5).
    الباك إند يعدّلها بناءً على:
    - إذا أي فريم NEUTRAL = -1 نجمة
    - إذا أي فريم يعاكس الإشارة = -2 نجمة
    - الحد الأدنى = 1 نجمة
    - 5 نجوم فقط إذا الثلاث فريمات متوافقة 100%
    """
    signal   = safe_get(data, "signal", "")
    bias_5m  = safe_get(data, "bias_5m", "")
    bias_15m = safe_get(data, "bias_15m", "")
    
    try:
        pine_stars = int(safe_get(data, "stars", "1"))
    except:
        pine_stars = 1
    
    adjusted = pine_stars
    
    # تحقق من 5m
    if signal == "CALL":
        if "Neutral" in bias_5m:
            adjusted -= 1
            logger.info(f"[Stars V7.1] -1 star: 5m is Neutral for CALL")
        elif "Bear" in bias_5m:
            adjusted -= 2
            logger.info(f"[Stars V7.1] -2 stars: 5m is Bear for CALL")
    elif signal == "PUT":
        if "Neutral" in bias_5m:
            adjusted -= 1
            logger.info(f"[Stars V7.1] -1 star: 5m is Neutral for PUT")
        elif "Bull" in bias_5m:
            adjusted -= 2
            logger.info(f"[Stars V7.1] -2 stars: 5m is Bull for PUT")
    
    # تحقق من 15m
    if signal == "CALL":
        if "Neutral" in bias_15m:
            adjusted -= 1
            logger.info(f"[Stars V7.1] -1 star: 15m is Neutral for CALL")
        elif "Bear" in bias_15m:
            adjusted -= 2
            logger.info(f"[Stars V7.1] -2 stars: 15m is Bear for CALL")
    elif signal == "PUT":
        if "Neutral" in bias_15m:
            adjusted -= 1
            logger.info(f"[Stars V7.1] -1 star: 15m is Neutral for PUT")
        elif "Bull" in bias_15m:
            adjusted -= 2
            logger.info(f"[Stars V7.1] -2 stars: 15m is Bull for PUT")
    
    # الحد الأدنى 1
    adjusted = max(1, adjusted)
    
    # 5 نجوم فقط إذا كل شيء متوافق
    if adjusted >= 5:
        all_aligned = True
        if signal == "CALL":
            if "Bull" not in bias_5m or "Bull" not in bias_15m:
                all_aligned = False
        elif signal == "PUT":
            if "Bear" not in bias_5m or "Bear" not in bias_15m:
                all_aligned = False
        if not all_aligned:
            adjusted = 4
    
    logger.info(f"[Stars V7.1] Pine={pine_stars} → Adjusted={adjusted} | signal={signal} | 5m={bias_5m} | 15m={bias_15m}")
    return adjusted

# ──────────────────────────────────────────────────────────────────────────────
# V7.1: Telegram Message Formatting — REDESIGNED
# ──────────────────────────────────────────────────────────────────────────────

MONTHS_AR = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
             "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]

def format_expiry_ar(expiry_str):
    try:
        d = datetime.strptime(expiry_str, "%Y-%m-%d")
        return f"{d.day} {MONTHS_AR[d.month - 1]}"
    except:
        return expiry_str


def format_trade_alert_v71(data, stars_adjusted, primary_opt=None, alt_opt=None, alpaca_msg=None):
    """
    V7.1: رسالة تداول عملية ومباشرة — للسكالبينج.
    لا تحليل أكاديمي. فقط: ادخل؟ وين الهدف؟ متى تطلع؟
    """
    signal   = safe_get(data, "signal", "?")
    price    = safe_get(data, "price", "?")
    bias     = safe_get(data, "bias", "--")
    cond     = safe_get(data, "cond", "--")
    vol      = safe_get(data, "vol", "--")
    bias_5m  = safe_get(data, "bias_5m", "--")
    bias_15m = safe_get(data, "bias_15m", "--")

    now_et    = get_et_now()
    timestamp = now_et.strftime("%I:%M %p")

    # ── أيقونات وقوة الإشارة ──
    if signal == "CALL":
        sig_icon = "🟢"
        direction = "CALL"
    else:
        sig_icon = "🔴"
        direction = "PUT"

    if stars_adjusted >= 4:
        strength = "قوي"
        action_text = "✅ هذا المكان حق الدخول — دخل الحين"
    elif stars_adjusted == 3:
        strength = "جيد"
        action_text = "✅ فرصة جيدة — دخل مع حذر"
    else:
        strength = "ضعيf"
        action_text = "⚠️ إشارة ضعيفة — الأفضل تنتظر"

    stars_display = "⭐" * stars_adjusted

    # ── حساب الأهداف ──
    try:
        p_float = float(price)
        if signal == "CALL":
            target_1 = p_float + 3.0
            target_2 = p_float + 5.0
            stop_level = p_float - 2.0
        else:
            target_1 = p_float - 3.0
            target_2 = p_float - 5.0
            stop_level = p_float + 2.0
    except:
        target_1 = 0
        target_2 = 0
        stop_level = 0

    # ── توافق الفريمات (مختصر) ──
    def frame_icon(bias_str, sig):
        if sig == "CALL":
            if "Bull" in bias_str: return "✅"
            elif "Bear" in bias_str: return "❌"
            else: return "➖"
        else:
            if "Bear" in bias_str: return "✅"
            elif "Bull" in bias_str: return "❌"
            else: return "➖"

    f5  = frame_icon(bias_5m, signal)
    f15 = frame_icon(bias_15m, signal)

    # ── الحجم ──
    if "Ultra" in vol or "Surge" in vol:
        vol_label = "قوي 💪"
    elif "Normal" in vol:
        vol_label = "طبيعي"
    else:
        vol_label = "ضعيف ⚠️"

    # ── بناء الرسالة ──
    msg = (
        f"{sig_icon} <b>{direction} {strength}</b> — TSLA ${price}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{action_text}\n"
        f"🎯 الهدف: ${target_1:.0f} → ${target_2:.0f}\n"
        f"🛑 وقف: إذا {'كسر' if signal == 'CALL' else 'فوق'} ${stop_level:.0f} اطلع فوراً\n"
        f"{stars_display} ({stars_adjusted}/5)\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📊 5m: {f5} | 15m: {f15} | حجم: {vol_label}\n"
    )

    # ── بيانات الأوبشن (مختصرة) ──
    if primary_opt and primary_opt.get("last_price", 0) > 0:
        p_strike = primary_opt["strike"]
        p_price  = primary_opt["last_price"]
        p_exp    = format_expiry_ar(primary_opt["expiry"])
        is_0dte  = primary_opt.get("is_0dte", False)
        dte_label = "0DTE ⚡" if is_0dte else p_exp

        msg += (
            f"💰 {signal} ${p_strike:.0f} ({dte_label}) @ ${p_price:.2f}\n"
        )

        if alt_opt and alt_opt.get("last_price", 0) > 0:
            a_strike = alt_opt["strike"]
            a_price  = alt_opt["last_price"]
            a_exp    = format_expiry_ar(alt_opt["expiry"])
            msg += f"🔄 بديل: ${a_strike:.0f} ({a_exp}) @ ${a_price:.2f}\n"

    # ── خريطة الانعكاسات ──
    try:
        p_float = float(price)
        signal_dir = "BULL" if signal == "CALL" else "BEAR"
        is_aligned, alignment_text, nearest = check_map_alignment(p_float, signal_dir)
        if alignment_text:
            msg += f"🗺️ {alignment_text}\n"
    except:
        pass

    msg += (
        f"━━━━━━━━━━━━━━━\n"
        f"⏱ إذا ما تحرك خلال 10 دق = اطلع\n"
    )

    # V7.1: Alpaca execution status
    if alpaca_msg:
        msg += f"🤖 ثاقب: {alpaca_msg}\n"

    msg += f"🕐 {timestamp} ET"

    return msg


def format_reversal_alert_v71(data, fib=None):
    """
    V7.1: رسالة انعكاس عملية — لا تقل "ادخل"، تقول "انتبه وانتظر التأكيد".
    """
    price    = safe_get(data, "price", "?")
    div_type = safe_get(data, "div_type", "?")
    tf       = safe_get(data, "timeframe", "5m")
    vol_ok   = safe_get(data, "vol_confirm", "false").lower() == "true"
    bias     = safe_get(data, "bias", "")
    bias_15m = safe_get(data, "bias_15m", "")

    now_et    = get_et_now()
    timestamp = now_et.strftime("%I:%M %p")

    if div_type == "BULL":
        rev_icon = "📈"
        direction_text = "صعود محتمل من هنا"
        confirm_text = "انتظر شمعة تأكيد خضراء + حجم"
    else:
        rev_icon = "📉"
        direction_text = "هبوط محتمل من هنا"
        confirm_text = "انتظر شمعة تأكيد حمراء + حجم"

    # ── الحجم ──
    if vol_ok:
        vol_text = "✅ حجم مؤكد — إشارة أقوى"
    else:
        vol_text = "⚠️ حجم ضعيف — انتبه"

    msg = (
        f"🔥 <b>منطقة انعكاس</b> — TSLA ${price}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"{rev_icon} {direction_text} (Divergence {tf})\n"
        f"⚠️ انتبه لهذا المكان — لا تدخل بعد\n"
        f"✅ {confirm_text}\n"
        f"📦 {vol_text}\n"
    )

    # ── تحذير ضد الترند ──
    if div_type == "BEAR" and ("Bull" in bias or "Bull" in bias_15m):
        msg += "🚨 تحذير: ضد الترند الصاعد — احتمال فشل عالي\n"
    elif div_type == "BULL" and ("Bear" in bias or "Bear" in bias_15m):
        msg += "🚨 تحذير: ضد الترند الهابط — احتمال فشل عالي\n"

    # ── خريطة الانعكاسات ──
    try:
        p_float = float(price)
        signal_dir = div_type
        is_aligned, alignment_text, nearest = check_map_alignment(p_float, signal_dir)
        if alignment_text:
            msg += f"🗺️ {alignment_text}\n"
    except:
        pass

    # ── V7.1: Fibonacci — FIXED direction ──
    if fib:
        try:
            p_float = float(price)
        except:
            p_float = 0

        if div_type == "BULL" and p_float > 0:
            # صعود: أظهر مستويات المقاومة فوق السعر الحالي
            targets = []
            for label, val in [("38.2%", fib["fib_382"]), ("50%", fib["fib_500"]), ("61.8%", fib["fib_618"])]:
                if val > p_float:
                    targets.append((label, val))
            # إذا لم نجد مستويات فوق، أظهر أقرب مستوى
            if not targets:
                targets = [("23.6%", fib["fib_236"])]
            
            if targets:
                target_str = " → ".join([f"${v:.0f}" for _, v in targets[:2]])
                msg += f"🎯 لو تأكد: هدف {target_str}\n"

        elif div_type == "BEAR" and p_float > 0:
            # هبوط: أظهر مستويات الدعم تحت السعر الحالي
            targets = []
            for label, val in [("61.8%", fib["fib_618"]), ("50%", fib["fib_500"]), ("38.2%", fib["fib_382"])]:
                if val < p_float:
                    targets.append((label, val))
            if not targets:
                targets = [("78.6%", fib["fib_786"])]
            
            if targets:
                target_str = " → ".join([f"${v:.0f}" for _, v in targets[:2]])
                msg += f"🎯 لو تأكد: هدف {target_str}\n"

    msg += (
        f"━━━━━━━━━━━━━━━\n"
        f"🕐 {timestamp} ET"
    )
    return msg


def format_opening_map_v71(price, d_high, d_low, fib, trend, liquidity):
    """V7.1: خريطة الصباح — عملية ومباشرة."""
    now_et = get_et_now()
    date_str = f"{now_et.day} {MONTHS_AR[now_et.month - 1]} {now_et.year}"
    timestamp = now_et.strftime("%I:%M %p")

    # V7.1: تحذير الجمعة
    friday_warning = ""
    if is_friday():
        friday_warning = "\n🚫 <b>اليوم جمعة — 0DTE = كازينو. لا تتداول!</b>\n"

    msg = (
        f"🗺️ <b>خريطة ثاقب — TSLA {date_str}</b>\n"
        f"━━━━━━━━━━━━━━━"
        f"{friday_warning}\n"
        f"💰 السعر: <code>${price:.2f}</code>\n"
        f"📈 الاتجاه: {trend} | السيولة: {liquidity}\n"
        f"━━━━━━━━━━━━━━━\n"
        f"📐 <b>مستويات اليوم:</b>\n"
        f"   🔴 ${fib['high']:.0f} (أعلى)\n"
        f"   ├ 38.2%: ${fib['fib_382']:.0f}\n"
        f"   ├ 50.0%: ${fib['fib_500']:.0f} ← المحور\n"
        f"   ├ 61.8%: ${fib['fib_618']:.0f} ← ذهبي\n"
        f"   🟢 ${fib['low']:.0f} (أدنى)\n"
        f"━━━━━━━━━━━━━━━\n"
    )

    # V7.1: GEX Levels (الأهم)
    gex = get_gex_levels()
    if gex and gex.get("gamma_flip"):
        gf = float(gex["gamma_flip"])
        cw = float(gex["call_wall"]) if gex.get("call_wall") else 0
        pw = float(gex["put_wall"]) if gex.get("put_wall") else 0
        regime = "إيجابي ✅" if price > gf else "سلبي ⚠️"
        msg += (
            f"🎯 <b>GEX (سمارت موني):</b>\n"
            f"   🟡 Gamma Flip: ${gf:.0f} ← المحور ({regime})\n"
            f"   🔴 Call Wall: ${cw:.0f} ← مقاومة (وقف هنا)\n"
            f"   🟢 Put Wall: ${pw:.0f} ← دعم (ارتداد محتمل)\n"
            f"━━━━━━━━━━━━━━━\n"
        )

    # V7.1: خريطة الانعكاسات المختصرة
    if reversal_map["built"] and reversal_map["levels"]:
        msg += "🗺️ <b>مستويات مهمة:</b>\n"
        count = 0
        for lvl in reversal_map["levels"]:
            if count >= 5:
                break
            icon = "🔴" if lvl["type"] == "resistance" else "🟢" if lvl["type"] == "support" else "🟡"
            msg += f"   {icon} ${lvl['price']:.0f} — {lvl['name']} ({lvl['strength']})\n"
            count += 1
        msg += "━━━━━━━━━━━━━━━\n"

    msg += (
        f"⏰ نافذة التداول: 9:30 AM – 3:30 PM ET\n"
        f"🤖 ثاقب V7.1 جاهز ✅\n"
        f"🕐 {timestamp} ET"
    )
    return msg


def format_choppy_warning_v71():
    """V7.1: تحذير السوق الشوبي — واضح ومباشر."""
    now_et = get_et_now()
    timestamp = now_et.strftime("%I:%M %p")
    
    msg = (
        f"⛔ <b>السوق شوبي — لا تتداول الحين</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"🔥 هالمكان حرق أعصاب وفلوس\n"
        f"انتظر رسالة CALL أو PUT جديدة\n"
        f"🕐 {timestamp} ET"
    )
    return msg


def format_rest_mode_v71():
    """V7.1: رسالة وضع الراحة."""
    now_et = get_et_now()
    timestamp = now_et.strftime("%I:%M %p")
    
    msg = (
        f"🛑 <b>ثاقب دخل وضع الراحة (30 دقيقة)</b>\n"
        f"━━━━━━━━━━━━━━━\n"
        f"السوق ما يتحرك — صمت 30 دقيقة لحمايتك\n"
        f"🕐 {timestamp} ET"
    )
    return msg


# ──────────────────────────────────────────────────────────────────────────────
# Telegram
# ──────────────────────────────────────────────────────────────────────────────

# ── Rate Limiter للتلقرام (V10.3) — يمنع الحجب بسبب الـ flood ──
_tg_last_sent = 0.0          # وقت آخر رسالة
_tg_msg_count = 0            # عدد الرسائل في النافذة الحالية
_tg_window_start = 0.0       # بداية النافذة الزمنية (60 ثانية)
_TG_MAX_PER_MINUTE = 20      # حد أقصى 20 رسالة/دقيقة (Telegram يسمح 30)
_TG_MIN_INTERVAL   = 1.5     # ثانية ونصف بين كل رسالة
_tg_lock = threading.Lock()

def send_telegram(message):
    global _tg_last_sent, _tg_msg_count, _tg_window_start
    url     = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id":                  TELEGRAM_CHAT_ID,
        "text":                     message,
        "parse_mode":               "HTML",
        "disable_web_page_preview": True
    }
    with _tg_lock:
        now = time.time()
        # إعادة ضبط العداد كل دقيقة
        if now - _tg_window_start >= 60:
            _tg_window_start = now
            _tg_msg_count = 0
        # فحص حد الدقيقة
        if _tg_msg_count >= _TG_MAX_PER_MINUTE:
            logger.warning(f"[TG Rate Limit] تجاوز الحد ({_TG_MAX_PER_MINUTE}/دقيقة) — تم تجاهل الرسالة")
            return False
        # فحص الفاصل الزمني بين الرسائل
        elapsed = now - _tg_last_sent
        if elapsed < _TG_MIN_INTERVAL:
            time.sleep(_TG_MIN_INTERVAL - elapsed)
        _tg_last_sent = time.time()
        _tg_msg_count += 1
    try:
        resp = http_requests.post(url, json=payload, timeout=10)
        if resp.status_code == 200:
            logger.info(f"Telegram: message sent [{_tg_msg_count}/{_TG_MAX_PER_MINUTE}]")
            return True
        elif resp.status_code == 429:
            # Telegram طلب انتظار
            retry_after = resp.json().get('parameters', {}).get('retry_after', 30)
            logger.warning(f"[TG] 429 Too Many Requests — انتظار {retry_after}s")
            time.sleep(retry_after)
            return False
        else:
            logger.error(f"Telegram error: {resp.status_code} -- {resp.text}")
            return False
    except Exception as e:
        logger.error(f"Telegram send failed: {e}")
        return False

# ──────────────────────────────────────────────────────────────────────────────
# Filters
# ──────────────────────────────────────────────────────────────────────────────

def check_data_quality(data):
    signal = safe_get(data, "signal", "")
    price  = safe_get(data, "price", "")
    if not signal or signal in ("", "--", "?", "UNKNOWN"):
        return False, "بيانات ناقصة (لا يوجد إشارة)"
    if not price or price in ("", "--", "?"):
        return False, "بيانات ناقصة (لا يوجد سعر)"
    return True, ""

def check_cooldown(data):
    global last_alert_time, last_alert_price, last_alert_signal
    global last_call_time, last_put_time
    global loss_cooldown_until

    now = time.time()

    if now < loss_cooldown_until:
        remaining = loss_cooldown_until - now
        return False, f"وضع الراحة مفعل — انتظر {remaining/60:.0f} دقيقة"

    elapsed = now - last_alert_time
    signal  = safe_get(data, "signal", "")
    current_price = safe_get(data, "price", "")

    if current_price == last_alert_price and signal == last_alert_signal and elapsed < COOLDOWN_MIN_GAP:
        return False, f"مكرر (نفس السعر {current_price} بفارق {elapsed:.0f}ث)"

    if elapsed < COOLDOWN_MIN_GAP:
        return False, f"سريع جداً ({elapsed:.0f}ث < {COOLDOWN_MIN_GAP}ث)"

    if signal == "CALL":
        if now - last_call_time < COOLDOWN_SECONDS_SIMILAR:
            remaining = COOLDOWN_SECONDS_SIMILAR - (now - last_call_time)
            return False, f"CALL cooldown — انتظر {remaining/60:.0f} دقيقة"
    elif signal == "PUT":
        if now - last_put_time < COOLDOWN_SECONDS_SIMILAR:
            remaining = COOLDOWN_SECONDS_SIMILAR - (now - last_put_time)
            return False, f"PUT cooldown — انتظر {remaining/60:.0f} دقيقة"

    return True, ""

def check_daily_limit(data=None):
    reset_daily_if_needed()
    if len(daily_alerts) >= MAX_DAILY_TRADE_ALERTS:
        return False, f"وصلت الحد اليومي ({MAX_DAILY_TRADE_ALERTS} تنبيهات)"
    return True, ""

def check_adx_filter(data):
    """V6.0: فلتر ADX إجباري — لا تداول في سوق جانبي."""
    cond = safe_get(data, "cond", "")
    if "Choppy" in cond or "Ranging" in cond or "choppy" in cond:
        return False, f"⛔ ADX منخفض (سوق جانبي) — لا دخول | cond={cond}"
    return True, ""

def check_gex_alignment(data):
    try:
        if not _FA_AVAILABLE:
            return True, ""
        signal = safe_get(data, "signal", "")
        price = float(safe_get(data, "price", "0"))
        if not signal or price <= 0:
            return True, ""
        aligned, reason = check_gex_alignment(signal, price)
        if not aligned:
            return False, f"📊 GEX غير متوافق: {reason}"
        return True, reason
    except Exception:
        return True, ""

def check_trend_alignment(data):
    """V7.0: فلتر توافق الإشارة مع اتجاه الترند."""
    signal   = safe_get(data, "signal", "")
    bias     = safe_get(data, "bias", "")
    bias_15m = safe_get(data, "bias_15m", "")
    stars    = safe_get(data, "stars", "1")
    
    try:
        stars_int = int(stars)
    except:
        stars_int = 1
    
    if stars_int >= 5:
        return True, ""
    
    if signal == "PUT" and "Bull" in bias and "Bull" in bias_15m:
        return False, f"⚠️ PUT ضد ترند صاعد قوي — مرفوض"
    
    if signal == "CALL" and "Bear" in bias and "Bear" in bias_15m:
        return False, f"⚠️ CALL ضد ترند هابط قوي — مرفوض"
    
    return True, ""

def check_15m_mandatory_filter(data):
    """
    V7.1: فلتر 15 دقيقة إجباري — BUG FIX #2.
    إذا 15m يعاكس الإشارة → يمنع التنفيذ على Alpaca.
    لكن يسمح بإرسال الرسالة مع تحذير.
    
    Returns: (allow_execution: bool, reason: str)
    """
    signal   = safe_get(data, "signal", "")
    bias_15m = safe_get(data, "bias_15m", "")
    
    if signal == "CALL" and "Bear" in bias_15m:
        return False, "15m هابط ضد CALL — لا تنفيذ"
    
    if signal == "PUT" and "Bull" in bias_15m:
        return False, "15m صاعد ضد PUT — لا تنفيذ"
    
    return True, ""

def check_friday_filter(data=None):
    """V7.1: لا تداول يوم الجمعة (0DTE = كازينو)."""
    if is_friday():
        return False, "🚫 يوم الجمعة — 0DTE = كازينو. لا تداول!"
    return True, ""


def apply_filters(data):
    """V7.1: فلاتر محسّنة — ADX + Friday + GEX + Trend + Cooldown + Daily."""
    for check in [check_data_quality, check_friday_filter, check_adx_filter, 
                  check_trend_alignment, check_gex_alignment, check_cooldown, check_daily_limit]:
        ok, reason = check(data)
        if not ok:
            return False, reason
    return True, ""

# ──────────────────────────────────────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/", methods=["GET"])
def home():
    reset_daily_if_needed()
    in_window, window_msg = is_trading_window()
    account = get_alpaca_account()
    return jsonify({
        "status":          "running",
        "service":         "Mosquito ثاقب V10.3 — Smart Trading Alert Bot",
        "version":         "10.3",
        "trading_window":  window_msg,
        "in_window":       in_window,
        "alpaca_balance":  f"${float(account.get('cash', 0)):,.2f}" if account else "N/A",
        "alpaca_status":   account.get("status", "N/A") if account else "disconnected",
        "active_positions": len(active_positions),
        "alerts_today":    len(daily_alerts),
        "reversals_today": daily_reversal_count,
        "remaining_trades": MAX_DAILY_TRADE_ALERTS - len(daily_alerts),
        "remaining_reversals": MAX_DAILY_REVERSAL_ALERTS - daily_reversal_count,
        "gex_available":   _gex_cache["data"] is not None,
        "filters":         "ADX + 15m Mandatory + Trend + GEX + Friday + Stars V7.1",
        "reversal_map":    f"{len(reversal_map['levels'])} levels" if reversal_map["built"] else "not built",
        "timestamp":       datetime.now(timezone.utc).isoformat()
    })


@app.route("/webhook", methods=["POST"])
def webhook():
    global last_alert_time, last_alert_price, last_alert_signal
    global last_call_time, last_put_time
    global last_reversal_time, daily_reversal_count
    global consecutive_no_move, loss_cooldown_until

    if WEBHOOK_SECRET:
        auth = request.headers.get("X-Webhook-Secret", "")
        if auth != WEBHOOK_SECRET:
            return jsonify({"error": "Unauthorized"}), 401

    try:
        data = request.get_json() if request.is_json else json.loads(request.data.decode("utf-8"))
    except Exception as e:
        logger.error(f"JSON parse error: {e}")
        return jsonify({"error": "Parse error"}), 400

    signal   = safe_get(data, "signal", "?")
    price    = safe_get(data, "price", "?")
    msg_type = safe_get(data, "type", "TRADE")

    logger.info(f"[V7.1] Received: signal={signal} | type={msg_type} | price=${price}")

    # ── تحديث مستويات S/R في reversal_detector ──
    try:
        wh_resistance = float(safe_get(data, "resistance", "0") or 0)
        wh_support    = float(safe_get(data, "support",    "0") or 0)
        if wh_resistance > 0 and wh_support > 0 and _RD_AVAILABLE:
            rd.update_levels_from_webhook(wh_resistance, wh_support)
    except Exception as _e:
        pass

    if price not in ("?", "--"):
        market_state["last_price"]   = price
        market_state["last_updated"] = datetime.now(timezone.utc).isoformat()

    # ── OPENING MAP ──────────────────────────────────────────────────────────
    if signal == "OPENING_MAP":
        try:
            live_price, d_high, d_low = get_tsla_day_data()
            if not live_price:
                live_price = float(safe_get(data, "price", "0") or 0)
                d_high = float(safe_get(data, "day_high", "0") or 0)
                d_low  = float(safe_get(data, "day_low",  "0") or 0)

            if d_high > 0: market_state["day_high"] = d_high
            if d_low  > 0: market_state["day_low"]  = d_low

            bias_str = safe_get(data, "bias", "")
            if "Bull" in bias_str or "bull" in bias_str:
                trend = "صاعد 📈"
            elif "Bear" in bias_str or "bear" in bias_str:
                trend = "هابط 📉"
            else:
                trend = "جانبي ↔️"

            vol_str = safe_get(data, "vol", "")
            if "Ultra" in vol_str or "Surge" in vol_str:
                liquidity = "عالية 💧💧"
            elif "Normal" in vol_str:
                liquidity = "طبيعية 💧"
            else:
                liquidity = "منخفضة ⚠️"

            # بناء خريطة الانعكاسات أولاً
            try:
                build_reversal_map()
            except:
                pass

            if d_high > 0 and d_low > 0 and d_high > d_low:
                fib = calc_fibonacci(d_high, d_low)
                tg_msg = format_opening_map_v71(live_price or float(price), d_high, d_low, fib, trend, liquidity)
            else:
                now_et = get_et_now()
                tg_msg = (
                    f"🗺️ <b>خريطة ثاقب — TSLA {now_et.strftime('%d/%m/%Y')}</b>\n"
                    f"━━━━━━━━━━━━━━━\n"
                    f"💰 السعر: <code>${price}</code>\n"
                    f"📈 الاتجاه: {trend}\n"
                    f"🤖 ثاقب V7.1 جاهز ✅\n"
                    f"🕐 {now_et.strftime('%I:%M %p')} ET"
                )

            tg_ok = send_telegram(tg_msg)
            
            return jsonify({"status": "opening_map_sent", "telegram": "sent" if tg_ok else "failed"}), 200

        except Exception as e:
            logger.error(f"Opening map error: {e}")
            return jsonify({"status": "error", "error": str(e)}), 200

    # ── REVERSAL ALERT ───────────────────────────────────────────────────────
    if signal == "REVERSAL_ALERT":
        now = time.time()
        
        # V7.1: حد يومي للانعكاسات
        reset_daily_if_needed()
        if daily_reversal_count >= MAX_DAILY_REVERSAL_ALERTS:
            return jsonify({"status": "blocked", "reason": f"reversal daily limit ({MAX_DAILY_REVERSAL_ALERTS})"}), 200
        
        if now - last_reversal_time < REVERSAL_COOLDOWN_SECS:
            remaining = REVERSAL_COOLDOWN_SECS - (now - last_reversal_time)
            return jsonify({"status": "blocked", "reason": f"reversal cooldown — {remaining/60:.0f} min"}), 200

        # V7.1: BUG FIX #3 — فلتر الحجم للانعكاسات
        vol_confirm = safe_get(data, "vol_confirm", "false").lower() == "true"
        if not vol_confirm:
            logger.info(f"[V7.1] Reversal blocked — volume not confirmed (vol_confirm=false)")
            return jsonify({"status": "blocked", "reason": "reversal volume not confirmed — V7.1 filter"}), 200

        try:
            ph = float(safe_get(data, "day_high", "0") or 0)
            pl = float(safe_get(data, "day_low",  "0") or 0)
            if ph > 0: market_state["day_high"] = ph
            if pl > 0: market_state["day_low"]  = pl
        except:
            pass

        d_high = market_state["day_high"]
        d_low  = market_state["day_low"]
        fib = calc_fibonacci(d_high, d_low) if (d_high > 0 and d_low > 0 and d_high > d_low) else None

        if not reversal_map["built"] or reversal_map["date"] != get_today():
            try:
                build_reversal_map()
            except:
                pass

        tg_msg = format_reversal_alert_v71(data, fib)
        tg_ok  = send_telegram(tg_msg)
        if tg_ok:
            last_reversal_time = now
            daily_reversal_count += 1

        return jsonify({"status": "reversal_sent", "telegram": "sent" if tg_ok else "failed"}), 200

    # ── VOL TRAP (chart only) ─────────────────────────────────────────────────
    if signal == "VOL_INTEL":
        return jsonify({"status": "vol_intel_chart_only"}), 200

    # ── LOSS COUNTER ─────────────────────────────────────────────────────────
    if last_alert_price and price not in ("?", "--"):
        try:
            prev_p = float(last_alert_price)
            curr_p = float(price)
            move_pct = abs(curr_p - prev_p) / prev_p
            if move_pct < 0.003:
                consecutive_no_move += 1
                if consecutive_no_move >= LOSS_COUNTER_MAX:
                    loss_cooldown_until = time.time() + LOSS_COOLDOWN_SECONDS
                    consecutive_no_move = 0
                    send_telegram(format_rest_mode_v71())
            else:
                consecutive_no_move = 0
        except:
            pass

    # ── TRADE SIGNAL (CALL / PUT) ────────────────────────────────────────────
    try:
        passed, rejection_reason = apply_filters(data)
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 200

    if not passed:
        blocked_today.append({
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "signal":    signal,
            "price":     price,
            "reason":    rejection_reason
        })
        
        # V7.1: إذا السبب هو ADX/Choppy، أرسل تحذير الشوبي
        if "ADX" in rejection_reason or "Choppy" in rejection_reason or "Ranging" in rejection_reason:
            send_telegram(format_choppy_warning_v71())
        
        return jsonify({"status": "blocked", "reason": rejection_reason}), 200

    # V7.1: بناء خريطة الانعكاسات
    if not reversal_map["built"] or reversal_map["date"] != get_today():
        try:
            build_reversal_map()
        except:
            pass

    # V7.1: إعادة حساب النجوم
    stars_adjusted = calculate_stars_v71(data)

    # Fetch Option Data
    primary_opt, alt_opt = get_best_option("TSLA", signal, price)

    # ── V7.1: فلتر 15m إجباري (BUG FIX #2) ──
    allow_execution, filter_15m_reason = check_15m_mandatory_filter(data)

    # ── ALPACA AUTO-EXECUTE ───────────────────────────────────────────────────
    alpaca_status_msg = None
    in_window, window_msg = is_trading_window()

    if in_window and allow_execution:
        if stars_adjusted >= MIN_STARS_TO_EXECUTE:
            exec_ok, order_id, exec_msg = execute_trade(signal, price, stars_adjusted, primary_opt)
            if exec_ok:
                alpaca_status_msg = "تم التنفيذ ✅"
                send_telegram(exec_msg)
                logger.info(f"Alpaca trade executed: {order_id}")
            else:
                alpaca_status_msg = f"فشل ❌ ({exec_msg})"
                logger.error(f"Alpaca execution failed: {exec_msg}")
        else:
            alpaca_status_msg = f"لم يُنفَّذ — {stars_adjusted}⭐ (أقل من {MIN_STARS_TO_EXECUTE})"
    elif in_window and not allow_execution:
        alpaca_status_msg = f"⚠️ لم يُنفَّذ — {filter_15m_reason}"
        logger.info(f"[V7.1] 15m filter blocked execution: {filter_15m_reason}")
    else:
        alpaca_status_msg = f"خارج نافذة التداول"

    # Format and send Telegram alert
    tg_msg = format_trade_alert_v71(data, stars_adjusted, primary_opt, alt_opt, alpaca_status_msg)
    tg_ok  = send_telegram(tg_msg)

    # Update state
    global _current_signal_type
    now = time.time()
    last_alert_time   = now
    last_alert_price  = price
    last_alert_signal = signal
    _current_signal_type = signal
    if signal == "CALL":
        last_call_time = now
    elif signal == "PUT":
        last_put_time = now

    entry = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "signal":    signal,
        "price":     price,
        "stars":     str(stars_adjusted),
        "stars_pine": safe_get(data, "stars", "?"),
        "alpaca":    alpaca_status_msg,
        "15m_filter": "passed" if allow_execution else filter_15m_reason
    }
    alert_history.insert(0, entry)
    if len(alert_history) > MAX_HISTORY:
        alert_history.pop()
    daily_alerts.append(entry)

    logger.info(f"[V7.1] SENT: {signal} @ ${price} | Stars: Pine={safe_get(data, 'stars', '?')}→Adj={stars_adjusted} | Alpaca: {alpaca_status_msg}")

    return jsonify({
        "status":   "processed",
        "telegram": "sent" if tg_ok else "failed",
        "alpaca":   alpaca_status_msg,
        "stars_adjusted": stars_adjusted
    }), 200


# ──────────────────────────────────────────────────────────────────────────────
# Test & Utility Endpoints
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/reversal_report", methods=["GET"])
def reversal_report():
    if not _RT_AVAILABLE:
        return jsonify({"error": "Reversal Tracker not available"}), 503
    report_text = rt.generate_performance_report()
    import csv, os
    rows = []
    tracker_file = "/home/ubuntu/reversal_tracker_log.csv"
    if os.path.isfile(tracker_file):
        with open(tracker_file, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
    return jsonify({
        "summary": report_text,
        "total_evaluated": len(rows),
        "signals": rows[-20:] if len(rows) > 20 else rows
    })

@app.route("/alpaca_status", methods=["GET"])
def alpaca_status():
    account = get_alpaca_account()
    in_window, window_msg = is_trading_window()
    if account:
        return jsonify({
            "connected":       True,
            "account_number":  account.get("account_number"),
            "cash":            f"${float(account.get('cash', 0)):,.2f}",
            "portfolio_value": f"${float(account.get('portfolio_value', 0)):,.2f}",
            "buying_power":    f"${float(account.get('buying_power', 0)):,.2f}",
            "status":          account.get("status"),
            "trading_window":  window_msg,
            "in_window":       in_window,
            "active_positions": len(active_positions)
        })
    return jsonify({"connected": False}), 500


@app.route("/positions", methods=["GET"])
def positions():
    return jsonify({
        "active_positions": len(active_positions),
        "positions": list(active_positions.values())
    })


@app.route("/test_trade", methods=["GET"])
def test_trade():
    """V7.1: اختبار رسالة تداول بالشكل الجديد."""
    test_data = {
        "signal": "CALL", "type": "TRADE_V5_6", "price": "425.00",
        "stars": "4", "bias": "Bullish", "vol": "Surge",
        "cond": "Trending (Clear)", "session": "Morning Momentum",
        "bias_5m": "Bull", "bias_15m": "Bull"
    }
    stars_adj = calculate_stars_v71(test_data)
    primary_opt = {
        "strike": 425.0, "expiry": get_today(),
        "last_price": 2.50, "tp": 3.50, "sl": 1.25, "is_0dte": True
    }
    tg_msg = format_trade_alert_v71(test_data, stars_adj, primary_opt, None, "اختبار V7.1 ✅")
    tg_ok = send_telegram(tg_msg)
    return jsonify({"status": "test_sent", "stars_adjusted": stars_adj, "telegram": "sent" if tg_ok else "failed"})


@app.route("/test_reversal", methods=["GET"])
def test_reversal():
    """V7.1: اختبار رسالة انعكاس بالشكل الجديد."""
    test_data = {
        "signal": "REVERSAL_ALERT", "price": "424.80",
        "div_type": "BULL", "timeframe": "5m",
        "vol_confirm": "true", "bias": "Bearish", "bias_15m": "Bear"
    }
    fib = calc_fibonacci(445.0, 420.0)
    tg_msg = format_reversal_alert_v71(test_data, fib)
    tg_ok = send_telegram(tg_msg)
    return jsonify({"status": "test_sent", "telegram": "sent" if tg_ok else "failed"})


@app.route("/test_choppy", methods=["GET"])
def test_choppy():
    """V7.1: اختبار رسالة الشوبي."""
    tg_msg = format_choppy_warning_v71()
    tg_ok = send_telegram(tg_msg)
    return jsonify({"status": "test_sent", "telegram": "sent" if tg_ok else "failed"})


@app.route("/test_opening_map", methods=["GET"])
def test_opening_map():
    """V7.1: اختبار خريطة الصباح."""
    test_price = 425.00
    test_high  = 445.08
    test_low   = 420.00
    fib = calc_fibonacci(test_high, test_low)
    tg_msg = format_opening_map_v71(test_price, test_high, test_low, fib, "هابط 📉", "طبيعية 💧")
    tg_ok = send_telegram(tg_msg)
    return jsonify({"status": "test_sent", "fib": fib, "telegram": "sent" if tg_ok else "failed"})


@app.route("/fetch_gex", methods=["GET"])
def fetch_gex_endpoint():
    """V7.1: سحب GEX يدوياً وإرسال الخريطة للتلقرام."""
    gex_data = fetch_flashalpha_gex()
    if gex_data:
        msg = format_gex_telegram()
        if msg:
            send_telegram(msg)
        # إعادة بناء خريطة الانعكاسات
        reversal_map["built"] = False
        build_reversal_map()
        return jsonify({
            "status": "ok",
            "gex": gex_data,
            "reversal_map_levels": len(reversal_map["levels"]),
            "fetches_today": _gex_cache["fetch_count_today"],
            "max_daily": _GEX_MAX_DAILY_FETCHES,
            "telegram": "sent"
        })
    return jsonify({"status": "error", "message": "Failed to fetch GEX data"}), 500


@app.route("/test_alpaca", methods=["GET"])
def test_alpaca_endpoint():
    account = get_alpaca_account()
    if account:
        in_window, window_msg = is_trading_window()
        msg = (
            f"🤖 <b>ثاقب V7.1 — Alpaca متصل ✅</b>\n"
            f"━━━━━━━━━━━━━━━\n"
            f"💰 الرصيد: ${float(account.get('cash', 0)):,.2f}\n"
            f"📊 Portfolio: ${float(account.get('portfolio_value', 0)):,.2f}\n"
            f"⚡ Buying Power: ${float(account.get('buying_power', 0)):,.2f}\n"
            f"━━━━━━━━━━━━━━━\n"
            f"⏰ {window_msg}\n"
            f"🕐 {get_et_now().strftime('%I:%M %p')} ET"
        )
        send_telegram(msg)
        return jsonify({"status": "ok", "cash": account.get("cash"), "window": window_msg})
    return jsonify({"status": "error", "message": "Alpaca connection failed"}), 500


@app.route("/reset", methods=["GET"])
def reset():
    global daily_alerts, daily_date, blocked_today, daily_reversal_count
    global last_call_time, last_put_time, last_alert_price, last_alert_signal, last_alert_time
    global last_reversal_time, consecutive_no_move, loss_cooldown_until
    daily_alerts        = []
    blocked_today       = []
    daily_reversal_count = 0
    daily_date          = get_today()
    last_call_time      = 0
    last_put_time       = 0
    last_alert_price    = ""
    last_alert_signal   = ""
    last_alert_time     = 0
    last_reversal_time  = 0
    consecutive_no_move = 0
    loss_cooldown_until = 0
    reversal_map["built"] = False
    reversal_map["levels"] = []
    return jsonify({"status": "reset", "message": "All counters cleared — V7.1 ثاقب"})


# ──────────────────────────────────────────────────────────────────────────────
# Position Status Route (for reversal_detector integration)
# ──────────────────────────────────────────────────────────────────────────────

@app.route("/position_status", methods=["POST"])
def position_status():
    if not _RD_AVAILABLE:
        return jsonify({"status": "error", "message": "reversal_detector not available"}), 503

    data = request.get_json(force=True, silent=True) or {}
    is_open = bool(data.get("open", False))
    rd.set_position_open(is_open)

    return jsonify({
        "status":        "ok",
        "position_open": is_open,
        "message":       f"Position {'opened' if is_open else 'closed'}"
    })


# ──────────────────────────────────────────────────────────────────────────────
# Background Workers
# ──────────────────────────────────────────────────────────────────────────────

def keep_alive_worker():
    while True:
        time.sleep(KEEP_ALIVE_INTERVAL)
        try:
            render_url = os.environ.get("RENDER_EXTERNAL_URL", "https://tsla-scalper-bot.onrender.com")
            http_requests.get(f"{render_url}/", timeout=15)
        except Exception as e:
            logger.debug(f"Keep-alive ping failed: {e}")


def gex_morning_worker():
    """
    V7.1: سحب GEX مرتين باليوم:
    1) 9:25 AM ET — قبل السوق (خريطة أساسية)
    2) 10:00 AM ET — بعد نص ساعة (تحديث بعد تحرك السوق)
    """
    _morning_sent = False
    _update_sent = False
    _last_date = ""
    
    while True:
        try:
            now = get_et_now()
            today = now.strftime("%Y-%m-%d")
            
            # Reset flags for new day
            if _last_date != today:
                _morning_sent = False
                _update_sent = False
                _last_date = today
            
            # Skip weekends
            if now.weekday() >= 5:
                time.sleep(300)
                continue
            
            # Fetch 1: 9:25 AM ET — خريطة الصباح
            if not _morning_sent and now.hour == 9 and 24 <= now.minute <= 30:
                gex_data = fetch_flashalpha_gex()
                if gex_data:
                    msg = format_gex_telegram()
                    if msg:
                        send_telegram(msg)
                        logger.info("[GEX] Morning map sent to Telegram (Fetch 1/2)")
                    # بناء خريطة الانعكاسات بعد سحب GEX
                    reversal_map["built"] = False
                    build_reversal_map()
                    # V9.4: تحديث PE Engine بمستويات GEX
                    if _V8_AVAILABLE:
                        pe_levels = {
                            "gamma_flip": float(gex_data.get("gamma_flip") or 0),
                            "call_wall":  float(gex_data.get("call_wall") or 0),
                            "put_wall":   float(gex_data.get("put_wall") or 0),
                            "max_gamma":  float(gex_data.get("zero_dte_magnet") or gex_data.get("max_positive_gamma") or 0),
                        }
                        update_gex_levels(pe_levels)
                        logger.info(f"[PE] GEX levels auto-updated from morning fetch: {pe_levels}")
                _morning_sent = True
                time.sleep(1800)  # انتظر 30 دقيقة
                continue
            
            # Fetch 2: 10:00 AM ET — تحديث بعد تحرك السوق
            if not _update_sent and now.hour == 10 and 0 <= now.minute <= 5:
                gex_data = fetch_flashalpha_gex()
                if gex_data:
                    msg = format_gex_telegram()
                    if msg:
                        update_msg = "🔄 <b>تحديث GEX — بعد نص ساعة</b>\n" + msg[msg.index("━"):]
                        send_telegram(update_msg)
                        logger.info("[GEX] 30-min update sent to Telegram (Fetch 2/2)")
                    # إعادة بناء الخريطة بالبيانات المحدثة
                    reversal_map["built"] = False
                    build_reversal_map()
                    # V9.4: تحديث PE Engine بمستويات GEX المحدثة
                    if _V8_AVAILABLE:
                        pe_levels = {
                            "gamma_flip": float(gex_data.get("gamma_flip") or 0),
                            "call_wall":  float(gex_data.get("call_wall") or 0),
                            "put_wall":   float(gex_data.get("put_wall") or 0),
                            "max_gamma":  float(gex_data.get("zero_dte_magnet") or gex_data.get("max_positive_gamma") or 0),
                        }
                        update_gex_levels(pe_levels)
                        logger.info(f"[PE] GEX levels auto-updated from 30-min fetch: {pe_levels}")
                _update_sent = True
                time.sleep(3600)
                continue
            
            time.sleep(30)
        except Exception as e:
            logger.error(f"[GEX] Morning worker error: {e}")
            time.sleep(300)


# ──────────────────────────────────────────────────────────────────────────────
# V8 Options Scalper Integration
# ──────────────────────────────────────────────────────────────────────────────
try:
    from options_scalper import (
        start_scalper, stop_scalper, get_scalper_status,
        set_reversal_map_ref, send_telegram as v8_send_telegram,
        execute_manual_itm, close_manual_itm, get_manual_status,
        add_journal_entry, update_journal_entry, get_journal_entries,
        save_journal_image, get_journal_stats,
        start_pe_engine, stop_pe_engine, get_pe_status,
        update_cheddar_flow, update_gex_levels
    )
    _V8_AVAILABLE = True
    logger.info("V9 Options Scalper module loaded ✅")
except ImportError as e:
    _V8_AVAILABLE = False
    logger.warning(f"V9 Options Scalper not available: {e}")

# V8 API Endpoints
@app.route('/v8/status', methods=['GET'])
def v8_status():
    if not _V8_AVAILABLE:
        return jsonify({"error": "V8 not available"}), 503
    return jsonify(get_scalper_status())

@app.route('/v8/stop', methods=['POST'])
def v8_stop():
    if not _V8_AVAILABLE:
        return jsonify({"error": "V8 not available"}), 503
    stop_scalper()
    return jsonify({"status": "stopping"})

@app.route('/v8/start', methods=['POST'])
def v8_start():
    if not _V8_AVAILABLE:
        return jsonify({"error": "V8 not available"}), 503
    if reversal_map.get("built"):
        set_reversal_map_ref(reversal_map)
    start_scalper()
    return jsonify({"status": "started"})


# ──────────────────────────────────────────────────────────────────────────────
# V9 Manual ITM Scalper — الخط 2 اليدوي
# ──────────────────────────────────────────────────────────────────────────────

# متغير لحفظ آخر نتيجة صفقة مغلقة
_last_manual_result = {"pnl": 0, "reason": "", "entry_price": 0, "exit_price": 0}

@app.route('/manual', methods=['GET'])
def manual_page():
    """Serve the manual ITM scalper web page."""
    return render_template('manual.html')


@app.route('/manual/status', methods=['GET'])
def manual_status():
    """Return current status for the manual scalper UI."""
    if not _V8_AVAILABLE:
        return jsonify({"error": "V9 not available"}), 503
    
    status = get_manual_status()
    # إضافة آخر نتيجة لو ما فيه صفقة مفتوحة
    if not status["has_position"]:
        status["last_pnl"] = _last_manual_result["pnl"]
        status["last_reason"] = _last_manual_result["reason"]
    return jsonify(status)


@app.route('/manual/buy', methods=['POST'])
def manual_buy():
    """Execute manual ITM buy order."""
    if not _V8_AVAILABLE:
        return jsonify({"success": False, "error": "V9 not available"}), 503
    
    global _active_journal_id
    data = request.get_json(force=True, silent=True) or {}
    option_type = data.get("type", "").lower()
    pre_trade   = data.get("pre_trade", {})

    if option_type not in ("call", "put"):
        return jsonify({"success": False, "error": "يجب تحديد call أو put"}), 400

    success, result = execute_manual_itm(option_type)

    if success:
        logger.info(f"[V9 Manual] Buy {option_type.upper()} executed: {result.get('symbol')} @ ${result.get('entry_price')}")
        # ── تحديث بيانات pre_trade على الـ record الذي أنشأه options_scalper (V10 FIX) ──
        # لا نُنشئ record جديد — options_scalper أنشأه بالفعل في execute_manual_itm()
        try:
            from options_scalper import update_journal_entry, get_manual_status
            # journal_id موجود في result من execute_manual_itm
            journal_id = result.get("journal_id")
            if journal_id and pre_trade:
                # تحديث بيانات pre_trade فقط على الـ record الموجود
                update_journal_entry(journal_id, {
                    "cheddar_flow":  pre_trade.get("cheddar_flow", ""),
                    "gamma_flip":    pre_trade.get("gamma_flip", ""),
                    "vwap_position": pre_trade.get("vwap_position", ""),
                    "or_position":   pre_trade.get("or_position", ""),
                    "entry_reason":  pre_trade.get("entry_reason", ""),
                })
            _active_journal_id = journal_id
        except Exception as je:
            logger.warning(f"[V10] Journal pre_trade update failed: {je}")
        # إرجاع journal_id للواجهة لربط رفع الصورة
        resp = {"success": True, **result}
        if _active_journal_id:
            resp["journal_id"] = _active_journal_id
        return jsonify(resp)
    else:
        logger.warning(f"[V9 Manual] Buy failed: {result.get('error')}")
        return jsonify({"success": False, **result}), 400


@app.route('/manual/sell', methods=['POST'])
def manual_sell():
    """Execute manual ITM sell order (stop loss button)."""
    global _last_manual_result
    
    if not _V8_AVAILABLE:
        return jsonify({"success": False, "error": "V9 not available"}), 503
    
    data = request.get_json(force=True, silent=True) or {}
    reason = data.get("reason", "manual_stop")
    
    success, result = close_manual_itm(reason=reason)
    
    if success:
        _last_manual_result = {
            "pnl": result.get("pnl", 0),
            "reason": result.get("reason", ""),
            "entry_price": result.get("entry_price", 0),
            "exit_price": result.get("exit_price", 0)
        }
        logger.info(f"[V9 Manual] Sell executed: PnL=${result.get('pnl', 0):+.2f} | Reason={reason}")
        resp = {"success": True, **result}
        if _active_journal_id:
            resp["journal_id"] = _active_journal_id
        return jsonify(resp)
    else:
        # إذا لم توجد صفقة مفتوحة — الصفقة أُغلقت تلقائياً بالفعل (TP/SL)
        error_msg = result.get("error", "")
        if "لا توجد صفقة" in error_msg:
            logger.info("[V9 Manual] Sell called but no open position — already auto-closed (TP/SL)")
            pnl = _last_manual_result.get("pnl", 0) if _last_manual_result else 0
            return jsonify({"success": True, "already_closed": True, "pnl": pnl, "reason": "auto"})
        return jsonify({"success": False, **result}), 400

# ──────────────────────────────────────────────────────────────────────────────
# V9.3 Trade Journal Routes — سجل الصفقات
# ──────────────────────────────────────────────────────────────────────────────

_active_journal_id = None  # ID الصفقة المفتوحة حالياً

@app.route('/journal', methods=['GET'])
def journal_page():
    """صفحة سجل الصفقات."""
    return render_template('journal.html')


@app.route('/journal/entries', methods=['GET'])
def journal_entries():
    """إرجاع قائمة الصفقات المسجلة."""
    if not _V8_AVAILABLE:
        return jsonify({"entries": [], "stats": {}})
    limit = request.args.get('limit', 50, type=int)
    entries = get_journal_entries(limit)
    stats   = get_journal_stats()
    return jsonify({"entries": entries, "stats": stats})


@app.route('/journal/add', methods=['POST'])
def journal_add():
    """إضافة صفقة جديدة للسجل (عند الدخول)."""
    global _active_journal_id
    if not _V8_AVAILABLE:
        return jsonify({"success": False}), 503
    data = request.get_json(force=True, silent=True) or {}
    # دعم pre_trade من manual.html
    pt = data.get("pre_trade", {})
    # حقول إلزامية + حقول الأسئلة الخمسة (V10)
    entry = {
        "status":        "open",
        "direction":     data.get("direction", ""),
        "symbol":        data.get("symbol", ""),
        "strike":        data.get("strike", 0),
        "entry_price":   data.get("entry_price", 0),
        "tsla_price":    data.get("tsla_price", 0),
        "delta":         data.get("delta", 0),
        "gex_position":  data.get("gex_position", ""),
        # ── أسئلة ما قبل الدخول (V10) ──
        "cheddar_flow":  data.get("cheddar_flow") or pt.get("cheddar_flow", ""),
        "gamma_flip":    data.get("gamma_flip")   or pt.get("gamma_flip", ""),
        "vwap_position": data.get("vwap_position") or pt.get("vwap_position", ""),
        "or_position":   data.get("or_position")  or pt.get("or_position", ""),
        "entry_reason":  data.get("entry_reason") or pt.get("entry_reason", ""),
        "notes":         data.get("notes", ""),
        "images":        [],
        "exit_price":    None,
        "pnl_dollar":    None,
        "exit_reason":   None,
        "exit_time":     None,
    }
    entry_id = add_journal_entry(entry)
    _active_journal_id = entry_id
    return jsonify({"success": True, "id": entry_id})


@app.route('/journal/close/<int:entry_id>', methods=['POST'])
def journal_close(entry_id):
    """تحديث صفقة بنتيجة الخروج."""
    global _active_journal_id
    if not _V8_AVAILABLE:
        return jsonify({"success": False}), 503
    data = request.get_json(force=True, silent=True) or {}
    updates = {
        "status":      "closed",
        "exit_price":  data.get("exit_price"),
        "pnl_dollar":  data.get("pnl_dollar"),
        "exit_reason": data.get("exit_reason", ""),
        "exit_time":   data.get("exit_time", ""),
    }
    update_journal_entry(entry_id, updates)
    if _active_journal_id == entry_id:
        _active_journal_id = None
    return jsonify({"success": True})


@app.route('/journal/image/<int:entry_id>', methods=['POST'])
def journal_upload_image(entry_id):
    """رفع صورة مرتبطة بصفقة."""
    if not _V8_AVAILABLE:
        return jsonify({"success": False}), 503
    if 'image' not in request.files:
        return jsonify({"success": False, "error": "لا يوجد ملf"}), 400
    file = request.files['image']
    if not file or file.filename == '':
        return jsonify({"success": False, "error": "ملف فارغ"}), 400
    # تحديد الامتداد
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'webp', 'gif'):
        ext = 'jpg'
    # label: before / after / chart (اختياري من الطلب)
    label = request.form.get('label', '')
    image_data = file.read()
    # V10.3: حفظ الصورة في SQLite كـ Base64 — دائمة حتى بعد Restart
    data_uri = save_journal_image(entry_id, image_data, ext, label=label)
    return jsonify({"success": True, "data_uri": data_uri})


@app.route('/journal/img/<filename>', methods=['GET'])
def journal_image(filename):
    """إرجاع صورة من مجلد السجل."""
    import os
    from flask import send_from_directory
    images_dir = os.path.join(os.path.dirname(__file__), 'static', 'journal_images')
    return send_from_directory(images_dir, filename)


@app.route('/journal/active_id', methods=['GET'])
def journal_active_id():
    """إرجاع ID الصفقة المفتوحة حالياً."""
    return jsonify({"active_id": _active_journal_id})


@app.route('/journal/export/csv', methods=['GET'])
def journal_export_csv():
    """تصدير جميع الصفقات كملف CSV للتحليل."""
    import csv
    import io
    from flask import Response
    try:
        from options_scalper import get_journal_entries
        entries = get_journal_entries(limit=10000)
        output = io.StringIO()
        fieldnames = [
            'id', 'created_at', 'status', 'direction', 'symbol', 'strike',
            'qty', 'entry_price', 'exit_price', 'pnl_dollar', 'pnl_pct',
            'exit_reason', 'closed_at', 'tsla_price', 'notes'
        ]
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        for e in entries:
            writer.writerow({k: e.get(k, '') for k in fieldnames})
        csv_data = output.getvalue()
        from datetime import datetime
        filename = f"tsla_trades_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        return Response(
            csv_data,
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/journal/export/excel', methods=['GET'])
def journal_export_excel():
    """تصدير جميع الصفقات كملف Excel (.xlsx) مع تنسيق كامل."""
    try:
        import io
        from flask import Response
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from options_scalper import get_journal_entries
        from datetime import datetime

        entries = get_journal_entries(limit=10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "سجل الصفقات"

        # ألوان
        green_fill  = PatternFill("solid", fgColor="1A6B3C")
        red_fill    = PatternFill("solid", fgColor="8B0000")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        alt_fill    = PatternFill("solid", fgColor="0D1F33")
        white_font  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        normal_font = Font(color="FFFFFF", name="Calibri", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="2A4A6B"),
            right=Side(style="thin", color="2A4A6B"),
            top=Side(style="thin", color="2A4A6B"),
            bottom=Side(style="thin", color="2A4A6B")
        )

        # عنوان التقرير
        ws.merge_cells('A1:N1')
        title_cell = ws['A1']
        title_cell.value = f"سجل صفقات TSLA — {datetime.now().strftime('%Y-%m-%d')}"
        title_cell.font = Font(color="FFD700", bold=True, size=14, name="Calibri")
        title_cell.fill = PatternFill("solid", fgColor="0A1628")
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30

        # رأس الجدول
        headers = [
            '#', 'التاريخ', 'النوع', 'الرمز', 'Strike',
            'دخول ($)', 'خروج ($)', 'P&L ($)', 'P&L (%)',
            'الحالة', 'سبب الإغلاق', 'TSLA', 'VWAP', 'ملاحظات'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 22

        # بيانات الصفقات
        for row_idx, e in enumerate(entries, 3):
            is_win = (e.get('pnl_dollar') or 0) > 0
            is_open = e.get('status') == 'open'
            row_fill = PatternFill("solid", fgColor="0D2B1A") if is_win else \
                       PatternFill("solid", fgColor="2B0D0D") if not is_open else \
                       alt_fill

            pnl = e.get('pnl_dollar')
            pnl_pct = e.get('pnl_pct')

            row_data = [
                e.get('id', row_idx - 2),
                e.get('created_at', '')[:16] if e.get('created_at') else '',
                e.get('direction', '').upper(),
                e.get('symbol', ''),
                e.get('strike', ''),
                e.get('entry_price', ''),
                e.get('exit_price', '') if not is_open else 'مفتوحة',
                f"+${pnl:.2f}" if pnl and pnl > 0 else (f"-${abs(pnl):.2f}" if pnl and pnl < 0 else ''),
                f"{pnl_pct:+.1f}%" if pnl_pct else '',
                'ربح ✅' if is_win else ('مفتوحة ⏳' if is_open else 'خسارة ❌'),
                e.get('close_reason', '') or e.get('exit_reason', ''),
                e.get('tsla_price', ''),
                e.get('vwap', ''),
                e.get('notes', '')
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.font = Font(
                    color="00FF88" if is_win else ("FF4444" if not is_open else "FFFFFF"),
                    name="Calibri", size=10,
                    bold=(col == 8)  # P&L bold
                )
                cell.alignment = center_align
                cell.border = thin_border
            ws.row_dimensions[row_idx].height = 18

        # عرض الأعمدة
        col_widths = [5, 18, 8, 22, 8, 10, 10, 10, 8, 12, 20, 8, 8, 25]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # صف الملخص
        if entries:
            closed = [e for e in entries if e.get('status') == 'closed']
            wins = [e for e in closed if (e.get('pnl_dollar') or 0) > 0]
            total_pnl = sum(e.get('pnl_dollar') or 0 for e in closed)
            win_rate = round(len(wins) / len(closed) * 100, 1) if closed else 0

            summary_row = len(entries) + 4
            ws.merge_cells(f'A{summary_row}:N{summary_row}')
            summary_cell = ws[f'A{summary_row}']
            summary_cell.value = (
                f"إجمالي: {len(entries)} صفقة | مغلقة: {len(closed)} | "
                f"نسبة الفوز: {win_rate}% | صافي P&L: ${total_pnl:+.2f}"
            )
            summary_cell.font = Font(color="FFD700", bold=True, size=11, name="Calibri")
            summary_cell.fill = PatternFill("solid", fgColor="0A1628")
            summary_cell.alignment = center_align

        ws.sheet_view.showGridLines = False

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"tsla_trades_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/journal/db/info', methods=['GET'])
def journal_db_info():
    """معلومات عن قاعدة بيانات Journal."""
    try:
        import os
        from options_scalper import _JOURNAL_DB, get_journal_stats
        db_path = _JOURNAL_DB
        db_size = os.path.getsize(db_path) if os.path.exists(db_path) else 0
        stats = get_journal_stats()
        return jsonify({
            'ok': True,
            'db_path': db_path,
            'db_size_kb': round(db_size / 1024, 2),
            'persistent': db_path.startswith('/data'),
            'total_trades': stats.get('total', 0),
            'closed_trades': stats.get('closed', 0),
            'win_rate': stats.get('win_rate', 0),
            'total_pnl': stats.get('total_pnl', 0),
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# V10 FIX — تنظيف الصفقات المفتوحة الوهمية (Ghost Open Trades Cleanup)
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/journal/cleanup-ghosts', methods=['POST'])
def journal_cleanup_ghosts():
    """
    V10 FIX: تنظيف الصفقات المفتوحة الوهمية.
    تُغلق كل صفقة بحالة 'open' ليس لها صفقة مغلقة مطابقة (نفس created_at).
    تُستدعى مرة واحدة لتنظيف البيانات القديمة.
    """
    try:
        from options_scalper import _JOURNAL_DB, _journal_db_lock, _get_db_conn
        import sqlite3
        
        cleaned = 0
        with _journal_db_lock:
            conn = _get_db_conn()
            # جلب كل الصفقات المفتوحة
            open_rows = conn.execute(
                "SELECT id, created_at, direction, entry_price FROM trades WHERE status = 'open' ORDER BY id"
            ).fetchall()
            
            for row in open_rows:
                row_id = row['id']
                created_at = row['created_at']
                
                # تحقق: هل يوجد صفقة مغلقة بنفس created_at (يعني هي مكررة)؟
                closed_match = conn.execute(
                    "SELECT id FROM trades WHERE status = 'closed' AND created_at = ? AND id != ?",
                    (created_at, row_id)
                ).fetchone()
                
                if closed_match:
                    # هذه الصفقة مكررة — احذفها
                    conn.execute("DELETE FROM trades WHERE id = ?", (row_id,))
                    cleaned += 1
                    logger.info(f"[V10 Cleanup] Deleted ghost open trade #{row_id} (duplicate of closed #{closed_match['id']})")
            
            conn.commit()
            conn.close()
        
        return jsonify({
            'ok': True,
            'cleaned': cleaned,
            'message': f'تم حذف {cleaned} صفقة مفتوحة وهمية'
        })
    except Exception as e:
        logger.error(f"[V10 Cleanup] Error: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# V9.4 — ITM Precision Entry Engine Routes
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/pe/status', methods=['GET'])
def pe_status():
    """إرجاع حالة محرك الدخول الدقيق."""
    if not _V8_AVAILABLE:
        return jsonify({"error": "PE not available"}), 503
    return jsonify(get_pe_status())


@app.route('/pe/cheddar', methods=['POST'])
def pe_update_cheddar():
    """
    تحديث نسبة CheddarFlow يدوياً.
    Body: {"call_pct": 65.5}
    """
    if not _V8_AVAILABLE:
        return jsonify({"error": "PE not available"}), 503
    data = request.get_json() or {}
    call_pct = data.get("call_pct")
    if call_pct is None:
        return jsonify({"error": "call_pct required"}), 400
    try:
        call_pct = float(call_pct)
        if not (0 <= call_pct <= 100):
            return jsonify({"error": "call_pct must be 0-100"}), 400
        update_cheddar_flow(call_pct)
        return jsonify({"ok": True, "call_pct": call_pct, "put_pct": 100 - call_pct})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/pe/gex', methods=['POST'])
def pe_update_gex():
    """
    تحديث مستويات GEX يدوياً.
    Body: {"gamma_flip": 410, "call_wall": 430, "put_wall": 390, "max_gamma": 415}
    """
    if not _V8_AVAILABLE:
        return jsonify({"error": "PE not available"}), 503
    data = request.get_json() or {}
    try:
        levels = {
            "gamma_flip": float(data.get("gamma_flip", 0)),
            "call_wall":  float(data.get("call_wall", 0)),
            "put_wall":   float(data.get("put_wall", 0)),
            "max_gamma":  float(data.get("max_gamma", 0)),
        }
        update_gex_levels(levels)
        return jsonify({"ok": True, "levels": levels})
    except Exception as e:
        return jsonify({"error": str(e)}), 500



# ──────────────────────────────────────────────────────────────────────────────
# Pair Trade Routes — XOM CALL + XLE PUT
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/pair/scan', methods=['GET'])
def pair_scan():
    """مسح العقود المتاحة بدون تنفيذ."""
    try:
        from options_scalper import scan_pair_contracts, get_stock_price, PAIR_XOM_CONTRACTS, PAIR_XLE_CONTRACTS
        xom_c, xle_c, expiry = scan_pair_contracts()
        if not xom_c or not xle_c:
            return jsonify({"ok": False, "error": "لم يُعثر على عقود ATM مناسبة في نافذة DTE 14-21"})
        xom_price = get_stock_price("XOM")
        xle_price = get_stock_price("XLE")
        xom_cost  = xom_c["mid"] * PAIR_XOM_CONTRACTS * 100
        xle_cost  = xle_c["mid"] * PAIR_XLE_CONTRACTS * 100
        total_cost   = round(xom_cost + xle_cost, 2)
        target_value = round(total_cost * 1.30, 2)
        return jsonify({
            "ok": True,
            "xom_price": xom_price, "xle_price": xle_price,
            "xom_symbol": xom_c["symbol"], "xom_strike": xom_c["strike"],
            "xom_mid": xom_c["mid"], "xom_qty": PAIR_XOM_CONTRACTS,
            "xle_symbol": xle_c["symbol"], "xle_strike": xle_c["strike"],
            "xle_mid": xle_c["mid"], "xle_qty": PAIR_XLE_CONTRACTS,
            "total_cost": total_cost, "target_value": target_value,
            "shared_expiry": expiry,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/pair/buy', methods=['POST'])
def pair_buy():
    """تنفيذ Pair Trade."""
    try:
        from options_scalper import execute_pair_trade
        ok, data = execute_pair_trade()
        return jsonify({"ok": ok, **data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/pair/sell', methods=['POST'])
def pair_sell():
    """إغلاق Pair Trade يدوياً."""
    try:
        from options_scalper import close_pair_trade
        ok, data = close_pair_trade(reason="manual")
        if not ok:
            return jsonify({"ok": False, **data})
        return jsonify({"ok": True, **data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/pair/status', methods=['GET'])
def pair_status():
    """حالة Pair Trade الحالية."""
    try:
        from options_scalper import get_pair_status
        data = get_pair_status()
        return jsonify({"ok": True, **data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ──────────────────────────────────────────────────────────────────────────────
# Reversal Warning Routes (V10.2)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/dashboard', methods=['GET'])
def dashboard():
    """لوحة مراقبة حية لجميع الـ endpoints."""
    return render_template('dashboard.html')

@app.route('/pyramid/status', methods=['GET'])
def pyramid_status():
    """حالة نظام Pyramid Auto Simulation V11.0."""
    try:
        from options_scalper import get_pyramid_status
        return jsonify(get_pyramid_status())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/strategy-b/status', methods=['GET'])
def strategy_b_status():
    """حالة Strategy B — VWAP Bounce 15M."""
    try:
        from options_scalper import get_strategy_b_status
        return jsonify(get_strategy_b_status())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/strategy-c/status', methods=['GET'])
def strategy_c_status():
    """حالة Strategy C — Opening Range Breakout."""
    try:
        from options_scalper import get_strategy_c_status
        return jsonify(get_strategy_c_status())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/reversal/status', methods=['GET'])
def reversal_warning_status():
    """حالة نظام تحذير انعكاس TSLA 5M."""
    try:
        from options_scalper import get_reversal_warning_status
        data = get_reversal_warning_status()
        return jsonify({"ok": True, **data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/mosquito/status', methods=['GET'])
def mosquito_status():
    """حالة نظام Mosquito Scanner — MACD متعدد الأطر الزمنية."""
    try:
        from options_scalper import get_mosquito_status
        return jsonify({"ok": True, **get_mosquito_status()})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ──────────────────────────────────────────────────────────────────────────────
# V11.0: ORB Smart Assistant — مساعد ORB الذكي (تنبيه فقط — لا تنفيذ تلقائي)
# ──────────────────────────────────────────────────────────────────────────────
try:
    from orb_assistant import start_orb_assistant, stop_orb_assistant, get_orb_status
    _ORB_AVAILABLE = True
    logger.info("ORB Smart Assistant module loaded ✅")
except ImportError as e:
    _ORB_AVAILABLE = False
    logger.warning(f"ORB Smart Assistant not available: {e}")

@app.route('/orb/status', methods=['GET'])
def orb_status():
    """حالة مساعد ORB الذكي."""
    try:
        if not _ORB_AVAILABLE:
            return jsonify({"ok": False, "error": "ORB module not available"}), 503
        return jsonify(get_orb_status())
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/orb/start', methods=['POST'])
def orb_start():
    """تشغيل مساعد ORB."""
    try:
        if not _ORB_AVAILABLE:
            return jsonify({"ok": False, "error": "ORB module not available"}), 503
        result = start_orb_assistant()
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route('/orb/stop', methods=['POST'])
def orb_stop():
    """إيقاف مساعد ORB."""
    try:
        if not _ORB_AVAILABLE:
            return jsonify({"ok": False, "error": "ORB module not available"}), 503
        result = stop_orb_assistant()
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ──────────────────────────────────────────────────────────────────────────────
# Startup
# ──────────────────────────────────────────────────────────────────────────────


def _market_briefing_worker():
    """
    Worker thread: يرسل Market Briefing كل 15 دقيقة
    يبدأ بعد 30 دقيقة من افتتاح السوق (10:00 AM ET)
    يتوقف عند 3:30 PM ET
    """
    from options_scalper import generate_market_briefing
    logger.info("[Briefing] Market Briefing worker started")

    # انتظر حتى الساعة 10:00 AM ET أول رسالة
    while True:
        try:
            now_et = datetime.now(timezone.utc) - timedelta(hours=4)
            h, m = now_et.hour, now_et.minute

            # نافذة التشغيل: 10:00 AM — 3:30 PM ET
            in_window = (h == 10 and m >= 0) or (10 < h < 15) or (h == 15 and m <= 30)

            if in_window:
                # توليد الرسالة
                msg = generate_market_briefing()
                if msg:
                    send_telegram(msg)
                    logger.info(f"[Briefing] Sent at {now_et.strftime('%H:%M ET')}")
                else:
                    logger.warning("[Briefing] generate_market_briefing returned None")

                # انتظر 15 دقيقة
                time.sleep(15 * 60)
            else:
                # خارج النافذة — انتظر دقيقة
                time.sleep(60)
        except Exception as e:
            logger.error(f"[Briefing] Worker error: {e}")
            time.sleep(60)


import os as _os
_threads_started = False
_threads_lock = threading.Lock()

def _start_background_threads():
    global _threads_started
    with _threads_lock:
        if _threads_started:
            return
        _threads_started = True
    logger.info(f"[Startup] PID={_os.getpid()} — starting background threads")
    threading.Thread(target=keep_alive_worker, daemon=True).start()
    threading.Thread(target=monitor_positions, daemon=True).start()
    logger.info("Background threads started (keep_alive + monitor_positions) ✅")
    if _RD_AVAILABLE:
        rd.start_reversal_detector()
        logger.info("Reversal Detector started ✅")
    if _RT_AVAILABLE:
        rt.start_reversal_tracker()
        logger.info("Reversal Tracker started ✅")
    if _AO_AVAILABLE:
        ao.start_options_feed(
            signal_type_fn   = _get_signal_type,
            current_price_fn = _get_tsla_price_for_options,
            update_callback  = _on_option_data_received
        )
        logger.info("Alpaca Options Feed started ✅")
    # V7.1: FlashAlpha GEX Direct API — always start
    threading.Thread(target=gex_morning_worker, daemon=True).start()
    logger.info("FlashAlpha GEX morning worker started (Direct API) ✅")
    # V10.2: TSLA 5M Reversal Warning System
    try:
        from options_scalper import start_reversal_warning
        start_reversal_warning()
        logger.info("TSLA 5M Reversal Warning System started ✅ 🔔")
    except Exception as _rw_err:
        logger.warning(f"Reversal Warning System not started: {_rw_err}")
    # Pair Trade: استعادة الصفقة المفتوحة إذا كانت موجودة
    try:
        from options_scalper import _load_pair_state
        _load_pair_state()
        logger.info("Pair Trade state loaded ✅")
    except Exception as _pair_load_err:
        logger.warning(f"Pair Trade state load failed: {_pair_load_err}")
    # V13.0: Strategy D — Mosquito Trend (ATM + Reinforcement)
    try:
        from options_scalper import start_mosquito
        result = start_mosquito()
        logger.info(f"V13.0 Strategy D (Mosquito Trend) started ✅ 🦟 result={result}")
    except Exception as _mq_err:
        logger.error(f"Strategy D (Mosquito) FAILED to start: {_mq_err}", exc_info=True)
    # V8: Options Scalper — autonomous trading engine
    if _V8_AVAILABLE:
        if reversal_map.get("built"):
            set_reversal_map_ref(reversal_map)
        start_scalper()
        logger.info("V8 Options Scalper Engine started ✅ 🚀")
        # V9.4: ITM Precision Entry Engine
        start_pe_engine()
        logger.info("V9.4 ITM Precision Entry Engine started ✅ 🎯")
    # V12.0: Strategy A — True Pyramid (DISABLED — WR=33%, P&L=-$831)
    if PYRAMID_AUTO_ENABLED:
        try:
            from options_scalper import start_pyramid_auto
            result = start_pyramid_auto()
            logger.info(f"V12.0 Pyramid Auto started ✅ 🦋 result={result}")
        except Exception as _pyr_err:
            logger.error(f"Pyramid Auto FAILED to start: {_pyr_err}", exc_info=True)
    else:
        logger.info("[V12.0 Pyramid Auto] DISABLED — PYRAMID_AUTO_ENABLED=False ⛔")
    # V11.1: Strategy B — VWAP Bounce 15M
    try:
        from options_scalper import start_strategy_b
        result = start_strategy_b()
        logger.info(f"V11.1 Strategy B (VWAP Bounce 15M) started ✅ 🎯 result={result}")
    except Exception as _stb_err:
        logger.error(f"Strategy B FAILED to start: {_stb_err}", exc_info=True)
    # V11.2: Strategy C — Opening Range Breakout
    try:
        from options_scalper import start_strategy_c
        result = start_strategy_c()
        logger.info(f"V11.2 Strategy C (ORB) started ✅ 📈 result={result}")
    except Exception as _stc_err:
        logger.error(f"Strategy C FAILED to start: {_stc_err}", exc_info=True)
    # V10.3: Market Briefing — رسالة تلقرام كل 15 دقيقة
    threading.Thread(target=_market_briefing_worker, daemon=True).start()
    logger.info("Market Briefing worker started ✅ 📊")
    # V11.0: ORB Smart Assistant — مساعد ORB الذكي (تنبيه فقط)
    if _ORB_AVAILABLE:
        try:
            result = start_orb_assistant()
            logger.info(f"V11.0 ORB Smart Assistant started ✅ 🧠 result={result}")
        except Exception as _orb_err:
            logger.error(f"ORB Smart Assistant FAILED to start: {_orb_err}", exc_info=True)
# # ── Auto-restart strategies on every request (handles Render sleep/wake) ──────
@app.before_request
def _ensure_strategies_alive():
    """يتحقق من حالة الاستراتيجيات ويُعيد تشغيلها إذا ماتت (بعد نوم Render)"""
    try:
        from options_scalper import (
            get_pyramid_status, start_pyramid_auto,
            get_strategy_b_status, start_strategy_b,
            get_strategy_c_status, start_strategy_c,
            get_mosquito_status, start_mosquito,
            _pyr_state, _pyr_lock,
            _stb_state, _stb_lock,
            _stc_state, _stc_lock,
            _std_state, _std_lock
        )
        import threading as _th
        # Pyramid (DISABLED)
        if PYRAMID_AUTO_ENABLED:
            pyr_status = get_pyramid_status()
            if not pyr_status.get("running"):
                with _pyr_lock:
                    _pyr_state["running"] = False
                start_pyramid_auto()
                logger.info("[AutoRestart] Pyramid V12 restarted ✅")
        # Strategy B
        stb_status = get_strategy_b_status()
        if not stb_status.get("running"):
            with _stb_lock:
                _stb_state["running"] = False
            start_strategy_b()
            logger.info("[AutoRestart] Strategy B restarted ✅")
        # Strategy C
        stc_status = get_strategy_c_status()
        if not stc_status.get("running"):
            with _stc_lock:
                _stc_state["running"] = False
            start_strategy_c()
            logger.info("[AutoRestart] Strategy C restarted ✅")
        # Strategy D (Mosquito)
        std_status = get_mosquito_status()
        if not std_status.get("running"):
            with _std_lock:
                _std_state["running"] = False
            start_mosquito()
            logger.info("[AutoRestart] Strategy D (Mosquito) restarted ✅")
    except Exception as _e:
        logger.error(f"[AutoRestart] error: {_e}", exc_info=True)
    # ORB Smart Assistant auto-restart
    try:
        if _ORB_AVAILABLE:
            orb_st = get_orb_status()
            if not orb_st.get("running"):
                start_orb_assistant()
                logger.info("[AutoRestart] ORB Smart Assistant restarted ✅ 🧠")
    except Exception as _orb_e:
        logger.error(f"[AutoRestart] ORB error: {_orb_e}")

# ──────────────────────────────────────────────────────────────────────────────
# V15.0: Dashboard APIs — Volume Levels, Reversal Gauges, Reversal Zones
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/api/volume-levels', methods=['GET'])
def api_volume_levels():
    """
    V15.0: جلب مستويات السيولة لـ 5 فريمات زمنية.
    يُرجع: لكل فريم — الحجم الحالي، المتوسط، النسبة (ratio)، المستوى (1-5)
    المستويات: 1=ضعيفة جداً، 2=ضعيفة، 3=متوسطة، 4=عالية، 5=عالية جداً
    """
    try:
        from options_scalper import get_tsla_bars as _get_bars
    except ImportError:
        _get_bars = None

    def _fetch_bars(timeframe, limit):
        """جلب شمعات TSLA من Alpaca."""
        try:
            if _get_bars:
                return _get_bars(timeframe, limit)
        except Exception:
            pass
        try:
            r = _retry_session.get(
                "https://data.alpaca.markets/v2/stocks/TSLA/bars",
                headers=alpaca_headers(),
                params={"timeframe": timeframe, "limit": limit, "feed": "iex", "sort": "asc"},
                timeout=10
            )
            if r.status_code == 200:
                return r.json().get("bars", [])
        except Exception as e:
            logger.error(f"[VolAPI] Bars error ({timeframe}): {e}")
        return []

    def _classify_volume(ratio):
        """تصنيف السيولة بناءً على النسبة مقارنة بالمتوسط."""
        if ratio is None:
            return 3, "متوسطة"
        if ratio < 0.5:
            return 1, "ضعيفة جداً"
        elif ratio < 0.8:
            return 2, "ضعيفة"
        elif ratio < 1.3:
            return 3, "متوسطة"
        elif ratio < 2.0:
            return 4, "عالية"
        else:
            return 5, "عالية جداً"

    def _calc_frame(timeframe, limit, lookback):
        """حساب بيانات السيولة لفريم زمني معين."""
        bars = _fetch_bars(timeframe, limit)
        if not bars or len(bars) < lookback + 1:
            return {"level": 3, "label": "متوسطة", "ratio": None, "current_vol": 0, "avg_vol": 0}
        volumes = [int(b["v"]) for b in bars]
        current_vol = volumes[-1]
        avg_vol = sum(volumes[-lookback-1:-1]) / lookback if lookback > 0 else current_vol
        ratio = round(current_vol / avg_vol, 2) if avg_vol > 0 else None
        level, label = _classify_volume(ratio)
        return {
            "level": level,
            "label": label,
            "ratio": ratio,
            "current_vol": current_vol,
            "avg_vol": round(avg_vol)
        }

    result = {
        "1H":  _calc_frame("1Hour",  30, 20),
        "15M": _calc_frame("15Min", 30, 20),
        "5M":  _calc_frame("5Min",  40, 20),
        "3M":  _calc_frame("3Min",  40, 20),
        "1M":  _calc_frame("1Min",  40, 20),
        "ok": True
    }
    return jsonify(result)


@app.route('/api/reversal-gauges', methods=['GET'])
def api_reversal_gauges():
    """
    V15.0: عدادات الانعكاس على فريم 5M.
    يُرجع: CALL/PUT Potential (0-100)، OBV Slope، Volume Reversal، Per Trade Gauge
    """
    try:
        from options_scalper import get_tsla_bars as _get_bars, _rw_rsi, _rw_ema, _rw_obv, _rw_macd_hist
    except ImportError:
        return jsonify({"ok": False, "error": "options_scalper not available"})

    try:
        bars = _get_bars("5Min", 60)
        if not bars or len(bars) < 30:
            return jsonify({
                "ok": True,
                "call_potential": 50, "put_potential": 50,
                "obv_gauge": 50, "vol_reversal": 50, "per_trade": 50,
                "direction": "NEUTRAL", "rsi": None, "macd_hist": None
            })

        closes  = [float(b["c"]) for b in bars]
        volumes = [int(b["v"]) for b in bars]

        # RSI
        rsi = _rw_rsi(closes, 14)

        # MACD Histogram
        macd_curr, macd_prev = _rw_macd_hist(closes)

        # OBV
        obv_list = _rw_obv(closes, volumes)
        obv_slope = 0
        if len(obv_list) >= 6:
            obv_slope = obv_list[-1] - obv_list[-6]

        # Volume Ratio (آخر شمعة vs متوسط 20)
        avg_vol_20 = sum(volumes[-21:-1]) / 20 if len(volumes) >= 21 else sum(volumes) / len(volumes)
        vol_ratio = volumes[-1] / avg_vol_20 if avg_vol_20 > 0 else 1.0

        # EMA9
        ema9 = _rw_ema(closes, 9)
        price = closes[-1]

        # ── حساب CALL Potential (0-100) ──────────────────────────────
        call_score = 50.0
        if rsi is not None:
            if rsi < 30:
                call_score += 20   # oversold
            elif rsi < 45:
                call_score += 10
            elif rsi > 70:
                call_score -= 20   # overbought
            elif rsi > 55:
                call_score -= 5

        if macd_curr is not None:
            if macd_curr > 0 and (macd_prev is None or macd_curr > macd_prev):
                call_score += 15   # MACD صاعد
            elif macd_curr < 0:
                call_score -= 15

        if obv_slope > 0:
            call_score += 10
        elif obv_slope < 0:
            call_score -= 10

        if ema9 and price > ema9:
            call_score += 5
        elif ema9 and price < ema9:
            call_score -= 5

        call_score = max(0, min(100, call_score))
        put_score  = 100 - call_score

        # ── OBV Gauge (0-100) ─────────────────────────────────────────
        obv_max = max(abs(obv_slope), 1)
        obv_gauge = min(100, max(0, int(50 + (obv_slope / obv_max) * 50)))

        # ── Volume Reversal Gauge (0-100) ─────────────────────────────
        vol_gauge = min(100, max(0, int(min(vol_ratio, 3.0) / 3.0 * 100)))

        # ── Per Trade Gauge (0-100) — نسبة الربح المتوقع ──────────────
        per_trade = int(call_score) if call_score >= 50 else int(put_score)

        # ── الاتجاه العام ─────────────────────────────────────────────
        if call_score >= 60:
            direction = "CALL"
        elif put_score >= 60:
            direction = "PUT"
        else:
            direction = "NEUTRAL"

        return jsonify({
            "ok": True,
            "call_potential": round(call_score),
            "put_potential":  round(put_score),
            "obv_gauge":      obv_gauge,
            "vol_reversal":   vol_gauge,
            "per_trade":      per_trade,
            "direction":      direction,
            "rsi":            round(rsi, 1) if rsi else None,
            "macd_hist":      round(macd_curr, 4) if macd_curr else None,
            "obv_slope":      round(obv_slope),
            "vol_ratio":      round(vol_ratio, 2),
            "ema9":           round(ema9, 2) if ema9 else None,
            "price":          round(price, 2)
        })

    except Exception as e:
        logger.error(f"[GaugeAPI] Error: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)})


@app.route('/api/reversal-zones', methods=['GET'])
def api_reversal_zones():
    """
    V15.0: مناطق الانعكاس السعرية اليومية.
    يُرجع: قائمة مناطق سعرية (نطاق $2 لكل منطقة) + تنبيه إذا السعر دخل المنطقة.
    """
    try:
        # جلب السعر الحالي
        price = get_tsla_price_alpaca_snapshot()

        # بناء/استخدام خريطة الانعكاسات الموجودة
        rmap = build_reversal_map()
        levels = rmap.get("levels", [])

        # إذا لا توجد مستويات، استخدم Alpaca Snapshot
        if not levels:
            try:
                r = _retry_session.get(
                    "https://data.alpaca.markets/v2/stocks/TSLA/snapshot",
                    headers=alpaca_headers(), timeout=8
                )
                if r.status_code == 200:
                    snap = r.json()
                    daily = snap.get("dailyBar", {})
                    prev  = snap.get("prevDailyBar", {})
                    day_h = float(daily.get("h", 0))
                    day_l = float(daily.get("l", 0))
                    prev_h = float(prev.get("h", 0))
                    prev_l = float(prev.get("l", 0))
                    if day_h > 0:
                        levels = [
                            {"name": "Day High",  "price": day_h,  "type": "resistance", "source": "Daily", "strength": "قوية"},
                            {"name": "Day Low",   "price": day_l,  "type": "support",    "source": "Daily", "strength": "قوية"},
                            {"name": "Prev High", "price": prev_h, "type": "resistance", "source": "Daily", "strength": "متوسطة"},
                            {"name": "Prev Low",  "price": prev_l, "type": "support",    "source": "Daily", "strength": "متوسطة"},
                        ]
            except Exception as _e:
                logger.error(f"[ZonesAPI] Snapshot error: {_e}")

        # تحويل المستويات إلى مناطق سعرية ($2 لكل منطقة)
        zones = []
        import pytz
        et = pytz.timezone("America/New_York")
        now_et = datetime.now(et)
        weekday_names = {0: "الاثنين", 1: "الثلاثاء", 2: "الأربعاء", 3: "الخميس", 4: "الجمعة", 5: "السبت", 6: "الأحد"}
        # اليوم التالي
        next_day = now_et.weekday() + 1
        if next_day > 6:
            next_day = 0
        next_day_name = weekday_names.get(next_day, "غداً")
        today_name = weekday_names.get(now_et.weekday(), "اليوم")

        for lvl in levels:
            lvl_price = float(lvl.get("price", 0))
            if lvl_price <= 0:
                continue
            zone_low  = round(lvl_price - 1.0, 2)
            zone_high = round(lvl_price + 1.0, 2)
            # تنبيه: هل السعر الحالي داخل المنطقة؟
            in_zone = (price > 0) and (zone_low <= price <= zone_high)
            # مسافة السعر من المنطقة
            dist = round(abs(price - lvl_price), 2) if price > 0 else None
            near = (dist is not None) and (dist <= 2.0)
            zones.append({
                "name":       lvl.get("name", "Level"),
                "price":      lvl_price,
                "zone_low":   zone_low,
                "zone_high":  zone_high,
                "type":       lvl.get("type", "pivot"),
                "source":     lvl.get("source", "—"),
                "strength":   lvl.get("strength", "متوسطة"),
                "in_zone":    in_zone,
                "near_zone":  near,
                "distance":   dist,
                "day_label":  today_name,
                "next_day":   next_day_name,
                "alert":      in_zone or near
            })

        # ترتيب: المناطق الأقرب للسعر أولاً
        if price > 0:
            zones.sort(key=lambda z: abs(z["price"] - price))

        return jsonify({
            "ok":         True,
            "price":      round(price, 2) if price > 0 else None,
            "zones":      zones,
            "date":       rmap.get("date", get_today()),
            "today":      today_name,
            "next_day":   next_day_name,
            "zone_count": len(zones)
        })

    except Exception as e:
        logger.error(f"[ZonesAPI] Error: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)})


# ──────────────────────────────────────────────────────────────────────────────
# V10 — Cheddar Flow Vision Analyzer (تحليل صور السيولة بالذكاء الاصطناعي)
# ──────────────────────────────────────────────────────────────────────────────

OPENAI_API_KEY_FLOW = os.environ.get("OPENAI_API_KEY", "")

CHEDDAR_VISION_PROMPT = """أنت محلل خوارزمي لسيولة عقود الأوبشن.
من هذه الصورة لواجهة Cheddar Flow:

1. استخرج نسبة Call Flow % ونسبة Put Flow %.
2. اقرأ أول 3 أسطر فقط من جدول الصفقات واستخرج:
   - نوع العقد: Call أو Put
   - جهة التنفيذ: Ask أو Above أو Bid
   - نوع الطلب: Sweep أو Block

أعد البيانات بصيغة JSON فقط دون أي شرح:
{
  "call_pct": 72,
  "put_pct": 28,
  "rows": [
    {"type": "Call", "side": "Ask", "order": "Sweep"},
    {"type": "Call", "side": "Above", "order": "Sweep"},
    {"type": "Put", "side": "Bid", "order": "Block"}
  ]
}"""


def _analyze_cheddar_image(image_bytes: bytes, mime_type: str = "image/jpeg") -> dict:
    """
    ترسل الصورة إلى GPT-4o Vision وتسترجع JSON.
    """
    import base64
    import json as _json
    api_key = OPENAI_API_KEY_FLOW
    if not api_key:
        raise ValueError("لم يتم تعيين OPENAI_API_KEY في متغيرات البيئة")
    
    img_b64 = base64.b64encode(image_bytes).decode("utf-8")
    
    payload = {
        "model": "gpt-4o",
        "max_tokens": 500,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": CHEDDAR_VISION_PROMPT},
                    {"type": "image_url", "image_url": {
                        "url": f"data:{mime_type};base64,{img_b64}",
                        "detail": "high"
                    }}
                ]
            }
        ]
    }
    
    resp = http_requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json=payload,
        timeout=30
    )
    resp.raise_for_status()
    
    content = resp.json()["choices"][0]["message"]["content"].strip()
    # تنظيف النص واستخراج JSON
    if content.startswith("```"):
        content = content.split("```")[1]
        if content.startswith("json"):
            content = content[4:]
    return _json.loads(content.strip())


def _apply_flow_decision(data: dict) -> dict:
    """
    خوارزمية اتخاذ القرار بناءً على بيانات Cheddar Flow.
    تُرجع: {"signal": "CALL"|"PUT"|"WAIT", "reason": "...", "call_pct": ..., "put_pct": ...}
    """
    call_pct = float(data.get("call_pct", 50))
    put_pct  = float(data.get("put_pct",  50))
    rows     = data.get("rows", [])
    
    # حساب عدد Call Sweeps على Ask/Above
    call_sweeps = sum(
        1 for r in rows
        if str(r.get("type", "")).lower() == "call"
        and str(r.get("order", "")).lower() == "sweep"
        and str(r.get("side", "")).lower() in ("ask", "above")
    )
    # حساب عدد Put Sweeps على Ask/Above
    put_sweeps = sum(
        1 for r in rows
        if str(r.get("type", "")).lower() == "put"
        and str(r.get("order", "")).lower() == "sweep"
        and str(r.get("side", "")).lower() in ("ask", "above")
    )
    
    # فحص فخ السيولة: Call عالية لكن أول 3 صفوف تظهر Put قوية
    fakeout = (
        call_pct >= 60
        and put_sweeps >= 2
    )
    
    if fakeout:
        signal = "WAIT"
        reason = f"فخ سيولة — Call عالية ({call_pct:.0f}%) لكن {put_sweeps} Put Sweeps مفاجئة في أول 3 صفوf"
    elif call_pct > 65 and call_sweeps >= 2:
        signal = "CALL"
        reason = f"سيولة صاعدة شرسة — Call {call_pct:.0f}% | {call_sweeps} Call Sweeps على Ask/Above"
    elif put_pct > 65 and put_sweeps >= 2:
        signal = "PUT"
        reason = f"سيولة هابطة شرسة — Put {put_pct:.0f}% | {put_sweeps} Put Sweeps على Ask/Above"
    else:
        signal = "WAIT"
        reason = f"لا يوجد إشارة واضحة — Call {call_pct:.0f}% / Put {put_pct:.0f}% | Call Sweeps: {call_sweeps} | Put Sweeps: {put_sweeps}"
    
    return {
        "signal":     signal,
        "reason":     reason,
        "call_pct":   call_pct,
        "put_pct":    put_pct,
        "call_sweeps": call_sweeps,
        "put_sweeps":  put_sweeps,
        "rows":       rows
    }


def _format_flow_telegram(result: dict) -> str:
    """تنسيق رسالة تيليغرام لنتيجة تحليل Cheddar Flow."""
    signal = result["signal"]
    emoji  = {"CALL": "🚀", "PUT": "🩸", "WAIT": "🛑"}.get(signal, "❓")
    color  = {"CALL": "🟢", "PUT": "🔴", "WAIT": "🟡"}.get(signal, "⚪")
    
    rows_text = ""
    for i, r in enumerate(result.get("rows", []), 1):
        t = r.get("type", "?"); s = r.get("side", "?"); o = r.get("order", "?")
        icon = "🟢" if t.lower() == "call" else "🔴"
        rows_text += f"  {i}. {icon} {t} | {s} | {o}\n"
    
    return (
        f"{color} <b>Cheddar Flow — {signal} {emoji}</b>\n"
        f"───────────────\n"
        f"📊 السيولة الكلية:\n"
        f"  🟢 Call: <b>{result['call_pct']:.0f}%</b>  |  🔴 Put: <b>{result['put_pct']:.0f}%</b>\n"
        f"🔍 أول 3 صفوف:\n{rows_text}"
        f"🧠 القرار: {result['reason']}\n"
        f"⏰ {get_et_now().strftime('%I:%M %p')} ET"
    )


@app.route('/analyze-flow', methods=['POST'])
def analyze_flow():
    """
    V10 Cheddar Flow Vision Analyzer.
    استقبال صورة Cheddar Flow وتحليلها بـ GPT-4o Vision.
    الإرسال: POST multipart/form-data بمفتاح 'image'
    """
    try:
        # استقبال الصورة
        if 'image' not in request.files:
            return jsonify({'ok': False, 'error': 'يجب إرسال ملف باسم image'}), 400
        
        img_file  = request.files['image']
        img_bytes = img_file.read()
        mime_type = img_file.content_type or 'image/jpeg'
        
        if len(img_bytes) == 0:
            return jsonify({'ok': False, 'error': 'الصورة فارغة'}), 400
        
        # تحليل الصورة بالذكاء الاصطناعي
        logger.info(f"[FlowAnalyzer] Analyzing image ({len(img_bytes)//1024}KB)...")
        vision_data = _analyze_cheddar_image(img_bytes, mime_type)
        
        # تطبيق خوارزمية القرار
        result = _apply_flow_decision(vision_data)
        
        # إرسال التوصية على تيليغرام
        tg_msg = _format_flow_telegram(result)
        send_telegram(tg_msg)
        
        logger.info(f"[FlowAnalyzer] Signal: {result['signal']} | {result['reason']}")
        
        return jsonify({
            'ok':     True,
            'signal': result['signal'],
            'reason': result['reason'],
            'call_pct':    result['call_pct'],
            'put_pct':     result['put_pct'],
            'call_sweeps': result['call_sweeps'],
            'put_sweeps':  result['put_sweeps'],
            'rows':        result['rows']
        })
    
    except ValueError as ve:
        logger.error(f"[FlowAnalyzer] Config error: {ve}")
        return jsonify({'ok': False, 'error': str(ve)}), 503
    except Exception as e:
        logger.error(f"[FlowAnalyzer] Error: {e}", exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# API: صفقات مع الصور للتحليل الذكي
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/trades', methods=['GET'])
def api_trades():
    """
    يُرجع الصفقات مع الصور المرفقة.
    Query params:
      - date: تاريخ بصيغة YYYY-MM-DD (اختياري، افتراضي: اليوم ET)
      - limit: عدد الصفقات (اختياري، افتراضي: 50)
      - include_images: true/false (اختياري، افتراضي: true)
    """
    import pytz
    try:
        date_str   = request.args.get('date', '')
        limit      = int(request.args.get('limit', 50))
        inc_images = request.args.get('include_images', 'true').lower() == 'true'

        entries = get_journal_entries(limit=500)

        if date_str:
            entries = [e for e in entries if e.get('timestamp', '').startswith(date_str)]
        else:
            et    = pytz.timezone('America/New_York')
            today = datetime.now(et).strftime('%Y-%m-%d')
            entries = [e for e in entries if e.get('timestamp', '').startswith(today)]
            date_str = today

        entries = entries[:limit]

        if not inc_images:
            for e in entries:
                e.pop('images', None)

        closed   = [e for e in entries if e.get('status') == 'closed']
        wins     = [e for e in closed if (e.get('pnl') or 0) > 0]
        losses   = [e for e in closed if (e.get('pnl') or 0) < 0]
        total_pnl = sum(e.get('pnl', 0) or 0 for e in closed)

        return jsonify({
            'ok':   True,
            'date': date_str,
            'count': len(entries),
            'stats': {
                'total_pnl': round(total_pnl, 2),
                'wins':      len(wins),
                'losses':    len(losses),
                'win_rate':  round(len(wins) / len(closed) * 100, 1) if closed else 0,
                'avg_win':   round(sum(e.get('pnl', 0) or 0 for e in wins)   / len(wins),   2) if wins   else 0,
                'avg_loss':  round(sum(e.get('pnl', 0) or 0 for e in losses) / len(losses), 2) if losses else 0,
            },
            'trades': entries
        })
    except Exception as e:
        logger.error(f'[API/trades] Error: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/trades/<int:trade_id>', methods=['GET'])
def api_trade_detail(trade_id):
    """إرجاع تفاصيل صفقة واحدة مع صورها كاملة."""
    try:
        entries = get_journal_entries(limit=500)
        trade   = next((e for e in entries if e.get('id') == trade_id), None)
        if not trade:
            return jsonify({'ok': False, 'error': 'trade not found'}), 404
        return jsonify({'ok': True, 'trade': trade})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# API: تنزيل نسخة احتياطية من قاعدة البيانات
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/backup/db', methods=['GET'])
def api_backup_db():
    """
    تنزيل قاعدة البيانات كاملة كملف .db للنسخ الاحتياطي.
    مثال: GET /api/backup/db
    """
    try:
        from flask import send_file
        from options_scalper import _JOURNAL_DB
        import os
        if not os.path.exists(_JOURNAL_DB):
            return jsonify({'ok': False, 'error': 'DB not found'}), 404
        return send_file(
            _JOURNAL_DB,
            as_attachment=True,
            download_name='journal_backup.db',
            mimetype='application/octet-stream'
        )
    except Exception as e:
        logger.error(f'[API/backup] Error: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/backup/json', methods=['GET'])
def api_backup_json():
    """
    تصدير كل الصفقات والصور كـ JSON كامل للنسخ الاحتياطي.
    مثال: GET /api/backup/json
    """
    try:
        entries = get_journal_entries(limit=9999)
        from flask import Response
        import json as _json
        data = _json.dumps({
            'backup_time': datetime.now(timezone.utc).isoformat(),
            'total_trades': len(entries),
            'trades': entries
        }, ensure_ascii=False, indent=2)
        return Response(
            data,
            mimetype='application/json',
            headers={'Content-Disposition': 'attachment; filename=trades_backup.json'}
        )
    except Exception as e:
        logger.error(f'[API/backup/json] Error: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# API: تحليل AI للصفقة مع الصورة
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/analyze-trade', methods=['POST'])
def api_analyze_trade():
    """
    تحليل صفقة باستخدام GPT-4o Vision.
    Body JSON:
      - trade_id: رقم الصفقة (int)
      أو
      - direction: CALL/PUT
      - entry: سعر الدخول
      - exit_price: سعر الخروج
      - pnl: الربح/الخسارة
      - reason: سبب الدخول
      - image_b64: صورة الشارت base64 (data:image/...;base64,...)
    """
    try:
        import openai, os, base64, re

        data       = request.get_json(force=True) or {}
        trade_id   = data.get('trade_id')
        direction  = data.get('direction', '')
        entry      = data.get('entry', '')
        exit_price = data.get('exit_price', '')
        pnl        = data.get('pnl', '')
        reason     = data.get('reason', '')
        image_b64  = data.get('image_b64', '')

        # إذا أرسل trade_id فقط — نجلب بياناته من DB
        if trade_id and not image_b64:
            entries = get_journal_entries(limit=500)
            trade   = next((e for e in entries if e.get('id') == int(trade_id)), None)
            if trade:
                direction  = direction  or trade.get('direction', '')
                entry      = entry      or trade.get('entry_price', '') or trade.get('entry', '')
                exit_price = exit_price or trade.get('exit_price', '')
                pnl        = pnl        or trade.get('pnl_dollar', '') or trade.get('pnl', '')
                reason     = reason     or trade.get('entry_reason', '') or trade.get('reason', '')
                imgs       = trade.get('images', [])
                if imgs:
                    first = imgs[0]
                    if isinstance(first, dict):
                        image_b64 = first.get('data', '') or first.get('image', '')
                    else:
                        image_b64 = first

        # بناء الرسالة
        try:
            pnl_val = float(pnl or 0)
        except (ValueError, TypeError):
            pnl_val = 0
        pnl_str = f"+${pnl_val:.0f}" if pnl_val > 0 else f"-${abs(pnl_val):.0f}"
        trade_info = (
            f"نوع الصفقة: {direction}\n"
            f"دخول: ${entry} | خروج: ${exit_price}\n"
            f"نتيجة: {pnl_str}\n"
            f"سبب الدخول: {reason}"
        )

        system_prompt = (
            "أنت محلل تداول خبير في خيارات TSLA. "
            "مهمتك تحليل الصفقة والشارت وإعطاء تغذية راجعة سريعة ومفيدة بالعربية.\n"
            "رد بهذا الترتيب دائماً:\n"
            "✅ القرار: [\u0635حيح/\u062eاطئ] ولماذا\n"
            "📊 الشارت: ماذا يظهر (RSI, MACD, OBV, الترند)\n"
            "⚠️ الخطأ: ما الذي كان يجب تجنبه\n"
            "🎯 الدرس: جملة واحدة واضحة\n"
            "كن موجزاً ومباشراً."
        )

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": [
                {"type": "text", "text": f"بيانات الصفقة:\n{trade_info}"}
            ]}
        ]

        # إضافة الصورة إذا وجدت
        if image_b64:
            # تأكد من صحة الصيغة
            if not image_b64.startswith('data:'):
                image_b64 = 'data:image/jpeg;base64,' + image_b64
            messages[1]["content"].append({
                "type": "image_url",
                "image_url": {"url": image_b64, "detail": "high"}
            })

        api_key = os.environ.get('OPENAI_API_KEY', '')
        client  = openai.OpenAI(api_key=api_key)

        resp = client.chat.completions.create(
            model="gpt-4o",
            messages=messages,
            max_tokens=600,
            temperature=0.3
        )

        analysis = resp.choices[0].message.content.strip()

        return jsonify({
            'ok':       True,
            'trade_id': trade_id,
            'analysis': analysis
        })

    except Exception as e:
        logger.error(f'[API/analyze-trade] Error: {e}', exc_info=True)
        return jsonify({'ok': False, 'error': str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# TM Simulator — محاكي التداول الانضباطي
# ──────────────────────────────────────────────────────────────────────────────
import uuid as _uuid
from datetime import datetime as _dt

_tm_sessions = {}  # session_id -> session_data

def _tm_new_session():
    """إنشاء جلسة محاكاة جديدة."""
    return {
        "id": str(_uuid.uuid4())[:8],
        "balance": 3000.0,
        "initial_balance": 3000.0,
        "daily_pnl": 0.0,
        "trades": [],
        "open_trade": None,
        "circuit_broken": False,
        "consecutive_losses": 0,
        "max_streak": 0,
        "current_streak": 0,
        "created_at": _dt.utcnow().isoformat(),
        "trade_count": 0,
        "loans": [],
        "loan_balance": 0.0,
        "total_profit": 0.0,
    }

@app.route('/tm-simulator', methods=['GET'])
def tm_simulator_page():
    """صفحة محاكي TM الانضباطي."""
    return render_template('tm_simulator.html')

@app.route('/tm/price', methods=['GET'])
def tm_price():
    """سعر TSLA الحقيقي للمحاكي."""
    try:
        price = get_tsla_price_alpaca_snapshot()
        if not price or price <= 0:
            price = get_tsla_price_alpaca()
        return jsonify({"ok": True, "price": round(float(price), 2)})
    except Exception as e:
        return jsonify({"ok": False, "price": 0, "error": str(e)})

@app.route('/tm/bars', methods=['GET'])
def tm_bars():
    """شموع TSLA للمحاكي — 1Min و 5Min."""
    try:
        timeframe = request.args.get('tf', '5Min')
        limit = int(request.args.get('limit', 78))
        end_dt = _dt.utcnow()
        start_dt = end_dt - timedelta(hours=14)
        params = {
            "timeframe": timeframe,
            "start": start_dt.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "end": end_dt.strftime("%Y-%m-%dT%H:%M:%SZ"),
            "limit": limit,
            "feed": "iex"
        }
        r = http_requests.get(
            "https://data.alpaca.markets/v2/stocks/TSLA/bars",
            headers=alpaca_headers(),
            params=params,
            timeout=10
        )
        if r.status_code == 200:
            bars = r.json().get("bars", [])
            return jsonify({"ok": True, "bars": bars})
        return jsonify({"ok": False, "bars": [], "status": r.status_code})
    except Exception as e:
        return jsonify({"ok": False, "bars": [], "error": str(e)})

@app.route('/tm/session/new', methods=['POST'])
def tm_new_session():
    """إنشاء جلسة محاكاة جديدة."""
    session = _tm_new_session()
    _tm_sessions[session["id"]] = session
    return jsonify({"ok": True, "session": session})

@app.route('/tm/session/<sid>', methods=['GET'])
def tm_get_session(sid):
    """جلب حالة الجلسة."""
    session = _tm_sessions.get(sid)
    if not session:
        return jsonify({"ok": False, "error": "session not found"}), 404
    return jsonify({"ok": True, "session": session})

@app.route('/tm/trade/open', methods=['POST'])
def tm_open_trade():
    """فتح صفقة في المحاكي مع التحقق من قوانين TM."""
    data = request.get_json(force=True, silent=True) or {}
    sid = data.get("session_id")
    session = _tm_sessions.get(sid)
    if not session:
        return jsonify({"ok": False, "error": "session not found"}), 404

    # قانون Circuit Breaker
    if session["circuit_broken"]:
        return jsonify({"ok": False, "error": "🚫 Circuit Breaker مفعّل — خسارة -$500 اليوم. أوقف التداول.", "circuit": True})

    # قانون لا صفقة مفتوحة
    if session["open_trade"]:
        return jsonify({"ok": False, "error": "⚠️ عندك صفقة مفتوحة — أغلقها أولاً."})

    direction = data.get("direction", "").upper()
    entry_price = float(data.get("entry_price", 0))
    tp = float(data.get("tp", 0))
    sl = float(data.get("sl", 0))
    contracts = int(data.get("contracts", 1))
    current_tsla = float(data.get("tsla_price", 0))
    vwap = float(data.get("vwap", 0))

    # التحقق من الحقول الإجبارية
    if direction not in ("CALL", "PUT"):
        return jsonify({"ok": False, "error": "يجب تحديد CALL أو PUT"})
    if entry_price <= 0:
        return jsonify({"ok": False, "error": "سعر الدخول مطلوب"})
    if tp <= 0 or sl <= 0:
        return jsonify({"ok": False, "error": "⚠️ OCO إجباري — يجب تحديد TP و SL قبل الدخول"})

    # تحذير ذكي: CALL تحت VWAP أو PUT فوق VWAP
    warning = None
    if direction == "CALL" and vwap > 0 and current_tsla < vwap:
        warning = "⚠️ تنبيه: السعر تحت VWAP — أنت تلتقط سكيناً ساقطة! تأكد من إشارة انعكاس قوية."
    elif direction == "PUT" and vwap > 0 and current_tsla > vwap:
        warning = "⚠️ تنبيه: السعر فوق VWAP — تداول PUT في ترند صاعد. تأكد من إشارة انعكاس قوية."

    # فتح الصفقة
    trade = {
        "id": session["trade_count"] + 1,
        "direction": direction,
        "entry_price": entry_price,
        "tp": tp,
        "sl": sl,
        "contracts": contracts,
        "tsla_entry": current_tsla,
        "opened_at": _dt.utcnow().isoformat(),
        "status": "open",
        "pnl": 0.0,
    }
    session["open_trade"] = trade
    session["trade_count"] += 1
    return jsonify({"ok": True, "trade": trade, "warning": warning})

@app.route('/tm/trade/close', methods=['POST'])
def tm_close_trade():
    """إغلاق الصفقة المفتوحة في المحاكي."""
    data = request.get_json(force=True, silent=True) or {}
    sid = data.get("session_id")
    session = _tm_sessions.get(sid)
    if not session:
        return jsonify({"ok": False, "error": "session not found"}), 404

    trade = session.get("open_trade")
    if not trade:
        return jsonify({"ok": False, "error": "لا توجد صفقة مفتوحة"})

    exit_price = float(data.get("exit_price", trade["entry_price"]))
    reason = data.get("reason", "manual")
    contracts = trade["contracts"]

    # حساب P&L
    if trade["direction"] == "CALL":
        pnl = (exit_price - trade["entry_price"]) * 100 * contracts
    else:
        pnl = (trade["entry_price"] - exit_price) * 100 * contracts
    pnl = round(pnl, 2)

    trade["exit_price"] = exit_price
    trade["pnl"] = pnl
    trade["reason"] = reason
    trade["closed_at"] = _dt.utcnow().isoformat()
    trade["status"] = "closed"

    # تحديث الجلسة
    session["daily_pnl"] = round(session["daily_pnl"] + pnl, 2)
    session["balance"] = round(session["balance"] + pnl, 2)
    if pnl > 0:
        session["total_profit"] = round(session.get("total_profit", 0) + pnl, 2)
    session["trades"].append(dict(trade))
    session["open_trade"] = None

    # Streak
    if pnl > 0:
        session["consecutive_losses"] = 0
        session["current_streak"] += 1
        if session["current_streak"] > session["max_streak"]:
            session["max_streak"] = session["current_streak"]
    else:
        session["current_streak"] = 0
        session["consecutive_losses"] += 1

    # Circuit Breaker
    if session["daily_pnl"] <= -500:
        session["circuit_broken"] = True

    # تحذير 3 خسائر متتالية
    revenge_warning = None
    if session["consecutive_losses"] >= 3:
        revenge_warning = "🛑 3 خسائر متتالية — توقف 30 دقيقة. لا تتداول الآن. ابدأ من جديد، التاجر الحقيقي يتعافى."

    # رسالة Streak
    streak_msg = None
    if session["current_streak"] == 3:
        streak_msg = "🎯 أنت في المنطقة! 3 صفقات رابحة متتالية."
    elif session["current_streak"] > 3:
        streak_msg = f"🔥 سلسلة {session['current_streak']} رابحة! استمر بانضباط."

    return jsonify({
        "ok": True,
        "pnl": pnl,
        "daily_pnl": session["daily_pnl"],
        "balance": session["balance"],
        "circuit_broken": session["circuit_broken"],
        "consecutive_losses": session["consecutive_losses"],
        "current_streak": session["current_streak"],
        "max_streak": session["max_streak"],
        "revenge_warning": revenge_warning,
        "streak_msg": streak_msg,
        "trade": trade,
    })

@app.route('/tm/session/<sid>/reset', methods=['POST'])
def tm_reset_session(sid):
    """إعادة تعيين الجلسة — ممنوع إذا الأرباح فوق $200."""
    data = request.get_json(force=True, silent=True) or {}
    force = data.get("force", False)
    session = _tm_sessions.get(sid)
    if session and not force:
        if session.get("total_profit", 0) > 200:
            return jsonify({"ok": False, "error": "⚠️ ممنوع إعادة التعيين — أرباحك فوق $200. استمر!"})
    new_session = _tm_new_session()
    new_session["id"] = sid
    _tm_sessions[sid] = new_session
    return jsonify({"ok": True, "session": new_session})


@app.route('/tm/session/<sid>/loan', methods=['POST'])
def tm_take_loan(sid):
    """أخذ قرض — بنكي (5% فائدة) أو صديق (بدون فائدة)."""
    data = request.get_json(force=True, silent=True) or {}
    session = _tm_sessions.get(sid)
    if not session:
        return jsonify({"ok": False, "error": "session not found"}), 404

    loan_type = data.get("type", "bank")  # bank or friend
    amount = float(data.get("amount", 500))
    if amount <= 0 or amount > 1500:
        return jsonify({"ok": False, "error": "المبلغ يجب أن يكون بين $1 و $1500"})

    interest = 0.05 if loan_type == "bank" else 0.0
    repay_amount = round(amount * (1 + interest), 2)

    loan = {
        "type": loan_type,
        "amount": amount,
        "interest": interest,
        "repay_amount": repay_amount,
        "taken_at": _dt.utcnow().isoformat(),
        "repaid": False
    }
    session["loans"].append(loan)
    session["loan_balance"] = round(session.get("loan_balance", 0) + repay_amount, 2)
    session["balance"] = round(session["balance"] + amount, 2)

    label = "قرض بنكي" if loan_type == "bank" else "دعم صديق"
    return jsonify({
        "ok": True,
        "message": f"✅ {label}: +${amount:.0f} | المطلوب إرجاعه: ${repay_amount:.0f}",
        "session": session
    })


# ═══════════════════════════════════════════════════════════════════
# TM Simulator V12.2 — Options Chain + Boost
# ═══════════════════════════════════════════════════════════════════

@app.route('/tm/options', methods=['GET'])
def tm_options():
    """جلب عقود ATM و ITM لـ TSLA بناءً على السعر الحالي."""
    try:
        from options_scalper import get_options_chain, get_option_quote
        import datetime as _dtmod

        price = get_tsla_price_alpaca_snapshot() or get_tsla_price_alpaca()
        if not price or price <= 0:
            return jsonify({"ok": False, "error": "لا يمكن جلب سعر TSLA"})

        price = float(price)
        option_type = request.args.get("type", "call").lower()

        today = _dtmod.date.today()
        expiry = today.strftime("%Y-%m-%d")

        strike_min_atm = round(price - 3, 0)
        strike_max_atm = round(price + 3, 0)
        atm_contracts = get_options_chain(expiry, option_type, strike_min_atm, strike_max_atm)

        if not atm_contracts:
            for i in range(1, 8):
                expiry = (today + _dtmod.timedelta(days=i)).strftime("%Y-%m-%d")
                atm_contracts = get_options_chain(expiry, option_type, strike_min_atm, strike_max_atm)
                if atm_contracts:
                    break

        if option_type == "call":
            strike_min_itm = round(price - 8, 0)
            strike_max_itm = round(price - 3, 0)
        else:
            strike_min_itm = round(price + 3, 0)
            strike_max_itm = round(price + 8, 0)
        itm_contracts = get_options_chain(expiry, option_type, strike_min_itm, strike_max_itm)

        atm_result = None
        if atm_contracts:
            best = min(atm_contracts, key=lambda c: abs(float(c.get("strike_price", 0)) - price))
            symbol = best.get("symbol", "")
            quote = get_option_quote(symbol)
            atm_result = {
                "symbol": symbol,
                "strike": float(best.get("strike_price", 0)),
                "expiry": expiry,
                "bid": quote["bid"] if quote else 0,
                "ask": quote["ask"] if quote else 0,
                "mid": quote["mid"] if quote else 0,
                "delta": 0.50,
                "type": "ATM"
            }

        itm_result = None
        if itm_contracts:
            scored = []
            for c in itm_contracts:
                strike = float(c.get("strike_price", 0))
                if option_type == "call":
                    itm_amount = price - strike
                else:
                    itm_amount = strike - price
                if itm_amount >= 3:
                    approx_delta = min(0.90, 0.50 + itm_amount * 0.05)
                    scored.append((c, itm_amount, approx_delta))

            if scored:
                scored.sort(key=lambda x: abs(x[2] - 0.70))
                best_itm = scored[0]
                symbol = best_itm[0].get("symbol", "")
                quote = get_option_quote(symbol)
                itm_result = {
                    "symbol": symbol,
                    "strike": float(best_itm[0].get("strike_price", 0)),
                    "expiry": expiry,
                    "bid": quote["bid"] if quote else 0,
                    "ask": quote["ask"] if quote else 0,
                    "mid": quote["mid"] if quote else 0,
                    "delta": round(best_itm[2], 2),
                    "type": "ITM",
                    "itm_amount": round(best_itm[1], 2)
                }

        return jsonify({
            "ok": True,
            "price": round(price, 2),
            "expiry": expiry,
            "option_type": option_type,
            "atm": atm_result,
            "itm": itm_result
        })
    except Exception as e:
        logger.error(f"[TM Options] Error: {e}", exc_info=True)
        return jsonify({"ok": False, "error": str(e)})


@app.route('/tm/trade/boost', methods=['POST'])
def tm_boost_trade():
    """تعزيز الصفقة المفتوحة — مرة واحدة فقط + فقط إذا رابح +$50."""
    data = request.get_json(force=True, silent=True) or {}
    sid = data.get("session_id")
    session = _tm_sessions.get(sid)
    if not session:
        return jsonify({"ok": False, "error": "session not found"}), 404

    trade = session.get("open_trade")
    if not trade:
        return jsonify({"ok": False, "error": "لا توجد صفقة مفتوحة"})

    if trade.get("boosted"):
        return jsonify({"ok": False, "error": "⚠️ التعزيز مرة واحدة فقط في الصفقة"})

    current_pnl = float(data.get("current_pnl", 0))
    if current_pnl < 50:
        return jsonify({"ok": False, "error": f"⚠️ التعزيز فقط عند ربح +$50 أو أكثر (الحالي: ${current_pnl:.0f})"})

    trade["contracts"] += 1
    trade["boosted"] = True
    trade["boost_price"] = float(data.get("boost_price", trade["entry_price"]))
    trade["boost_at"] = _dt.utcnow().isoformat()

    old_qty = trade["contracts"] - 1
    new_avg = ((trade["entry_price"] * old_qty) + trade["boost_price"]) / trade["contracts"]
    trade["avg_entry"] = round(new_avg, 2)

    return jsonify({
        "ok": True,
        "message": f"✅ تعزيز ناجح! العقود: {trade['contracts']} | متوسط: ${new_avg:.2f}",
        "trade": trade
    })


# ══════════════════════════════════════════════════════════════════════════════
# TM Sniper Webhook — يستقبل إشارات TradingView ويرسلها لتلغرام
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/tm/webhook', methods=['POST'])
def tm_sniper_webhook():
    """Receive TM Sniper alerts from TradingView and forward to Telegram."""
    data = request.get_json(force=True, silent=True)
    if not data:
        # TradingView sometimes sends plain text in message body
        raw = request.get_data(as_text=True)
        if raw:
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                # Plain text message — wrap it
                data = {"message": raw}
        else:
            return jsonify({"ok": False, "error": "no data"}), 400

    action  = data.get("action", "UNKNOWN")
    message = data.get("message", "")
    price   = data.get("price", "")
    env     = data.get("env", "")

    if not message:
        message = f"TM Sniper Alert: {action} @ ${price}"

    # إرسال لتلغرام
    success = send_telegram(message)

    logger.info(f"[TM Webhook] {action} | ${price} | env={env} | sent={success}")

    return jsonify({
        "ok": True,
        "action": action,
        "price": price,
        "env": env,
        "telegram_sent": success
    })


# استدعاء مباشر عند بدء التشغيل
_start_background_threads()

if __name__ == "__main__":
    app.run(host=SERVER_HOST, port=SERVER_PORT, debug=False)
